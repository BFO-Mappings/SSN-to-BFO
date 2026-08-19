#!/usr/bin/env python3
"""Generate the authoritative SSN2BFO ontology from the COMS workbook."""

from __future__ import annotations

import argparse
import hashlib
import json
import re
import shlex
import shutil
import subprocess
import sys
import time
from collections import Counter, defaultdict
from dataclasses import dataclass, field
from datetime import datetime, timezone
from pathlib import Path
from typing import Iterable

from coms_row_identity import (
    CANONICALIZATION_VERSION,
    CanonicalRowAudit,
    CanonicalRowInput,
    ComsRowIdentityError,
    ExpressionNode as CanonicalExpressionNode,
    RowIdentityReference,
    RowLocation,
    build_row_audit,
    validate_row_id,
    validate_unique_authoritative_axioms,
    validate_unique_row_ids,
)
from product_dispositions import (
    DispositionDocument,
    DispositionRowInput,
    ProductDispositionError,
    RequiredInputHashes,
    axiom_input_from_canonical_row,
    build_disposition_document,
    serialize_disposition_document,
    validate_disposition_file,
)
from modular_products import (
    ModularReasoningResult,
    ModularProductError,
    ModularProductResult,
    ProductDispositionReconciliation,
    build_alignment_core,
    build_cco_extension,
    build_fixed_source_closure,
    build_fixed_validation_closure,
    build_strict_bfo_mapping,
    reconcile_product_axioms,
    render_authoritative_axiom_lines,
    select_product_axioms,
    serialize_modular_product,
    validate_alignment_core,
    validate_cco_extension,
    validate_strict_bfo_mapping,
)
from publication_metadata import (
    METADATA_PREFIXES,
    PublicationMetadata,
    PublicationMetadataError,
    release_project_imports,
    render_ontology_header_bytes,
    strip_emitted_ontology_header,
    load_metadata,
    validate_serialized_ontology_header,
)
from release_context import FormalReleaseContext, validate_formal_release_context

try:
    import openpyxl
except ModuleNotFoundError as exc:  # pragma: no cover - runtime dependency guard
    raise SystemExit("Missing dependency: openpyxl is required to inspect the COMS workbook.") from exc

try:
    from rdflib import BNode, Graph, Literal, Namespace, RDF, RDFS, OWL, URIRef
    from rdflib.collection import Collection
    from rdflib.compare import isomorphic
    from rdflib.namespace import XSD
except ModuleNotFoundError as exc:  # pragma: no cover - runtime dependency guard
    raise SystemExit("Missing dependency: rdflib is required to generate and validate RDF.") from exc


REPO_ROOT = Path(__file__).resolve().parents[1]
LEGACY_ONTOLOGY = REPO_ROOT / "legacy/SSN2BFO-pre-COMS.ttl"
IDENTITY_MODULE = REPO_ROOT / "tools/coms_row_identity.py"
DISPOSITION_MODULE = REPO_ROOT / "tools/product_dispositions.py"
MODULAR_PRODUCTS_MODULE = REPO_ROOT / "tools/modular_products.py"
PUBLICATION_METADATA = REPO_ROOT / "config/publication-metadata.toml"
GENERATED_NOTICE = (
    "# GENERATED FILE: produced from mappings/SSN2BFO-COMS.xlsx; "
    "do not edit SSN2BFO.ttl directly."
)

BASE_REQUIRED_COLUMNS = (
    "sssom:subject_id",
    "sssom:predicate_id",
    "coms:Target",
    "coms:Reasoning",
)
ROW_ID_HEADER = "coms:RowID"
REQUIRED_COLUMNS = BASE_REQUIRED_COLUMNS + (ROW_ID_HEADER,)

ONTOLOGY_IRI = URIRef("http://www.sks.ai/SSN2BFO/")
DIRECT_IMPORTS = (
    URIRef("http://www.w3.org/ns/ssn/"),
    URIRef("http://www.w3.org/ns/sosa/sampling/"),
    URIRef("http://www.w3.org/ns/ssn/systems/"),
    URIRef("https://www.commoncoreontologies.org/2024-11-06/CommonCoreOntologiesMerged"),
)

SOSA = Namespace("http://www.w3.org/ns/sosa/")
SSN = Namespace("http://www.w3.org/ns/ssn/")
SSN_SYSTEM = Namespace("http://www.w3.org/ns/ssn/systems/")
SAMPLING = Namespace("http://www.w3.org/ns/sosa/sampling/")
BFO = Namespace("http://purl.obolibrary.org/obo/")
CCO = Namespace("https://www.commoncoreontologies.org/")
COMS_COVERAGE = Namespace("http://www.sks.ai/SSN2BFO/coms/coverage#")

PREFIXES = {
    "bfo": str(BFO),
    "cco": str(CCO),
    "owl": str(OWL),
    "rdf": str(RDF),
    "rdfs": str(RDFS),
    "sampling": str(SAMPLING),
    "sosa": str(SOSA),
    "ssn": str(SSN),
    "ssn-system": str(SSN_SYSTEM),
}
ROOT_ONTOLOGY_DECLARATION_COUNT = 1
ROOT_IMPORT_COUNT = 4
ROOT_METADATA_ANNOTATION_COUNT = 7
ROOT_LOGICAL_TRIPLE_COUNT = 1102
ROOT_TOTAL_TRIPLE_COUNT = 1114
ROOT_FORMAL_TOTAL_TRIPLE_COUNT = 1117
ROOT_FIXED_CLOSURE_TRIPLE_COUNT = 15904
ROOT_FORMAL_FIXED_CLOSURE_TRIPLE_COUNT = 15907

PREFIX_FILES = {
    "bfo": Path("imports/cco.ttl"),
    "cco": Path("imports/cco.ttl"),
    "sampling": Path("imports/sosa-sampling.ttl"),
    "sosa": Path("imports/sosa.ttl"),
    "ssn": Path("imports/ssn.ttl"),
    "ssn-system": Path("imports/ssn-systems.ttl"),
}

SOURCE_IMPORTS = (
    Path("imports/sosa.ttl"),
    Path("imports/sosa-sampling.ttl"),
    Path("imports/ssn.ttl"),
    Path("imports/ssn-systems.ttl"),
)

BFO_VALIDATION_DEPENDENCY = Path("imports/cco.ttl")

CANDIDATE_CLOSURE_INPUTS = (
    Path("imports/cco.ttl"),
    Path("imports/sosa.ttl"),
    Path("imports/sosa-sampling.ttl"),
    Path("imports/ssn.ttl"),
    Path("imports/ssn-systems.ttl"),
)

ALLOWED_PREDICATES = {
    "rdfs:subClassOf": RDFS.subClassOf,
    "owl:equivalentClass": OWL.equivalentClass,
    "rdfs:subPropertyOf": RDFS.subPropertyOf,
    "owl:equivalentProperty": OWL.equivalentProperty,
    "owl:propertyChainAxiom": OWL.propertyChainAxiom,
    "rdfs:domain": RDFS.domain,
    "rdfs:range": RDFS.range,
}

CLASS_PREDICATES = {"rdfs:subClassOf", "owl:equivalentClass"}
OBJECT_PROPERTY_PREDICATES = {
    "rdfs:subPropertyOf",
    "owl:equivalentProperty",
    "owl:propertyChainAxiom",
}
DOMAIN_RANGE_PREDICATES = {"rdfs:domain", "rdfs:range"}
OBJECT_PROPERTY_SUBJECT_PREDICATES = OBJECT_PROPERTY_PREDICATES | DOMAIN_RANGE_PREDICATES
MAPPING_PREDICATES = CLASS_PREDICATES | OBJECT_PROPERTY_PREDICATES

SAMPLE_PROPERTY_SOURCE_DECLARATIONS = (
    (SOSA.isSampleOf, RDF.type, OWL.FunctionalProperty),
    (SOSA.hasSample, RDF.type, OWL.InverseFunctionalProperty),
)

# Diagnostic-only compatibility alias. Production and release validation must
# retain these pinned source declarations.
CLEANUP_TRIPLES = SAMPLE_PROPERTY_SOURCE_DECLARATIONS

DL_PROFILE_DECLARATION_COMPLETION = (
    (
        URIRef("http://purl.org/dc/terms/type"),
        RDF.type,
        OWL.AnnotationProperty,
    ),
    (
        URIRef("http://purl.org/dc/terms/issued"),
        RDF.type,
        OWL.AnnotationProperty,
    ),
    (
        URIRef("http://www.w3.org/ns/adms#status"),
        RDF.type,
        OWL.AnnotationProperty,
    ),
    (
        URIRef("http://www.w3.org/2004/02/skos/core#Concept"),
        RDF.type,
        OWL.Class,
    ),
)

TOKEN_RE = re.compile(
    r"\s*("
    r"\("
    r"|\)"
    r"|[A-Za-z][A-Za-z0-9_-]*:[A-Za-z_][A-Za-z0-9_-]*"
    r"|and\b"
    r"|or\b"
    r"|some\b"
    r"|o\b"
    r")",
)
UNSAT_RE = re.compile(r"unsatisfiable:\s+(\S+)")


@dataclass(frozen=True)
class WorkbookRow:
    sheet: str
    row_number: int
    subject_text: str
    predicate_text: str
    target_text: str
    reasoning_text: str
    stable_row_id: str
    mapping_status_text: str = ""

    @property
    def row_id(self) -> str:
        return f"{self.sheet}!{self.row_number}"

    @property
    def diagnostic_id(self) -> str:
        suffix = f" [{self.stable_row_id}]" if self.stable_row_id else ""
        return f"{self.row_id}{suffix}"

    @property
    def location(self) -> RowLocation:
        return RowLocation(self.sheet, self.row_number)

    @property
    def is_blank_mapping(self) -> bool:
        return bool(self.subject_text) and not self.predicate_text and not self.target_text

@dataclass(frozen=True)
class Resolution:
    token: str
    iri: URIRef
    kind: str
    method: str
    label: str = ""


@dataclass
class LabelResolutionRecord:
    token: str
    iri: URIRef
    kind: str
    method: str
    label: str
    rows: set[str] = field(default_factory=set)


@dataclass(frozen=True)
class Expr:
    kind: str
    iri: URIRef | None = None
    children: tuple["Expr", ...] = ()
    prop: URIRef | None = None
    filler: "Expr | None" = None


@dataclass
class ProcessedRow:
    row: WorkbookRow
    subject: URIRef
    subject_kind: str
    predicate: str
    target: str
    expr: Expr | None = None
    target_property: URIRef | None = None
    property_chain: tuple[URIRef, ...] = ()
    identity_audit: CanonicalRowAudit | None = None


@dataclass(frozen=True)
class NormalizedRow:
    row_id: str
    subject: str
    subject_kind: str
    predicate: str
    original_target: str
    normalized_target: str
    rdf_owl_form: str


@dataclass
class WorkbookStats:
    worksheets_read: list[str] = field(default_factory=list)
    rows_by_sheet: Counter[str] = field(default_factory=Counter)
    populated_rows_by_sheet: Counter[str] = field(default_factory=Counter)
    mapped_rows: int = 0
    blank_mapping_rows: int = 0
    class_mapping_rows: int = 0
    object_property_mapping_rows: int = 0
    domain_rows: int = 0
    range_rows: int = 0
    property_chain_rows: int = 0
    governed_row_id_count: int = 0
    unique_row_id_count: int = 0
    processed_row_count: int = 0
    identity_audit_row_count: int = 0
    identity_count_reconciliation_passed: bool = False
    identity_row_id_set_reconciliation_passed: bool = False
    identity_location_reconciliation_passed: bool = False

    @property
    def active_axiom_rows(self) -> int:
        return self.mapped_rows + self.domain_rows + self.range_rows


@dataclass
class HermitResult:
    graph_path: Path
    reasoned_path: Path
    generated_triple_count: int
    closure_triple_count: int
    return_code: int | None
    reasoned_output_produced: bool
    owl_nothing_count: int | None
    unsat_classes: list[URIRef]
    robot_output: str
    robot_path: str | None
    profile_checked: bool = False
    profile_graph_path: Path | None = None
    profile_triple_count: int | None = None
    profile_declaration_completion_count: int = 0
    profile_return_code: int | None = None
    profile_output: str = ""
    source_sample_declarations_retained: bool | None = None

    @property
    def passed(self) -> bool:
        profile_passed = (
            not self.profile_checked
            or (
                self.profile_return_code == 0
                and self.source_sample_declarations_retained is True
            )
        )
        return (
            profile_passed
            and self.return_code == 0
            and self.reasoned_output_produced
            and self.owl_nothing_count == 0
            and not self.unsat_classes
        )


@dataclass(frozen=True)
class RenderedOntologyProduct:
    product_key: str
    serialized_bytes: bytes
    ontology_declaration_triple_count: int
    import_triple_count: int
    metadata_annotation_count: int
    formal_metadata_annotation_count: int
    logical_triple_count: int
    total_triple_count: int
    governed_axiom_count: int
    sha256: str


@dataclass(frozen=True)
class FormalProductSet:
    context: FormalReleaseContext
    integrated: RenderedOntologyProduct
    alignment_core: ModularProductResult
    strict_bfo_mapping: ModularProductResult
    cco_extension: ModularProductResult
    reconciliations: tuple[ProductDispositionReconciliation, ...]


def robot_command_report_value(hermit: HermitResult | None) -> str:
    if hermit is None:
        return "not evaluated"
    if hermit.robot_path is None:
        return "`robot` (not found on `PATH`)"
    return "`robot` (resolved from `PATH`)"


@dataclass
class CoverageResult:
    source_terms: dict[URIRef, str]
    mapped_terms: set[URIRef]
    property_typing_terms: set[URIRef]
    listed_terms: set[URIRef]
    explicit_blank_terms: set[URIRef]
    spreadsheet_missing_subjects: set[str]
    query_source_count: int
    query_unmapped_count: int

    @property
    def mapped_classes(self) -> set[URIRef]:
        return {term for term in self.mapped_terms if self.source_terms.get(term) == "class"}

    @property
    def mapped_object_properties(self) -> set[URIRef]:
        return {term for term in self.mapped_terms if self.source_terms.get(term) == "object_property"}

    @property
    def unmapped_classes(self) -> set[URIRef]:
        return {
            term
            for term, kind in self.source_terms.items()
            if kind == "class" and term not in self.mapped_terms
        }

    @property
    def unmapped_object_properties(self) -> set[URIRef]:
        return {
            term
            for term, kind in self.source_terms.items()
            if kind == "object_property"
            and term not in self.mapped_terms
            and term not in self.property_typing_terms
        }

    @property
    def absent_terms(self) -> set[URIRef]:
        return set(self.source_terms) - self.listed_terms

    @property
    def property_typing_only_terms(self) -> set[URIRef]:
        return self.property_typing_terms - self.mapped_terms


@dataclass
class ComparisonResult:
    both: set[tuple[str, str, str, str]]
    coms_only: set[tuple[str, str, str, str]]
    legacy_only: set[tuple[str, str, str, str]]
    class_expression_differences: list[str]
    object_property_differences: list[str]
    property_chain_differences: list[str]
    domain_both: set[tuple[str, str]]
    domain_coms_only: set[tuple[str, str]]
    domain_legacy_only: set[tuple[str, str]]
    domain_differences: list[str]
    range_both: set[tuple[str, str]]
    range_coms_only: set[tuple[str, str]]
    range_legacy_only: set[tuple[str, str]]
    range_differences: list[str]

    @property
    def legacy_domain_range_absent_from_coms(self) -> set[tuple[str, str, str]]:
        domains = {
            (subject, str(RDFS.domain), target)
            for subject, target in self.domain_legacy_only
        }
        ranges = {
            (subject, str(RDFS.range), target)
            for subject, target in self.range_legacy_only
        }
        return domains | ranges


class GenerationError(Exception):
    pass


class ManchesterParser:
    def __init__(self, text: str, resolver: "Resolver", row_id: str):
        self.text = text
        self.resolver = resolver
        self.row_id = row_id
        self.tokens = self._tokenize(text)
        self.pos = 0

    def _tokenize(self, text: str) -> list[str]:
        tokens: list[str] = []
        cursor = 0
        while cursor < len(text):
            match = TOKEN_RE.match(text, cursor)
            if not match:
                excerpt = text[cursor : cursor + 30]
                raise GenerationError(f"{self.row_id}: malformed Manchester expression near {excerpt!r}")
            tokens.append(match.group(1))
            cursor = match.end()
        return tokens

    def parse(self) -> Expr:
        if not self.tokens:
            raise GenerationError(f"{self.row_id}: blank Manchester expression")
        expr = self._parse_or()
        if self._peek() is not None:
            raise GenerationError(f"{self.row_id}: unexpected token {self._peek()!r}")
        return expr

    def _peek(self) -> str | None:
        if self.pos >= len(self.tokens):
            return None
        return self.tokens[self.pos]

    def _consume(self, expected: str | None = None) -> str:
        token = self._peek()
        if token is None:
            raise GenerationError(f"{self.row_id}: unexpected end of expression")
        if expected is not None and token != expected:
            raise GenerationError(f"{self.row_id}: expected {expected!r}, found {token!r}")
        self.pos += 1
        return token

    def _parse_or(self) -> Expr:
        children = [self._parse_and()]
        while self._peek() == "or":
            self._consume("or")
            children.append(self._parse_and())
        if len(children) == 1:
            return children[0]
        return Expr("union", children=tuple(children))

    def _parse_and(self) -> Expr:
        children = [self._parse_primary()]
        while self._peek() == "and":
            self._consume("and")
            children.append(self._parse_primary())
        if len(children) == 1:
            return children[0]
        return Expr("intersection", children=tuple(children))

    def _parse_primary(self) -> Expr:
        token = self._peek()
        if token == "(":
            self._consume("(")
            expr = self._parse_or()
            self._consume(")")
            return expr
        if token is None:
            raise GenerationError(f"{self.row_id}: unexpected end of expression")
        if ":" not in token:
            raise GenerationError(f"{self.row_id}: expected CURIE or parenthesized expression, found {token!r}")
        curie = self._consume()
        if self._peek() == "some":
            self._consume("some")
            prop = self.resolver.resolve(curie, "object_property", self.row_id).iri
            filler = self._parse_primary()
            return Expr("some", prop=prop, filler=filler)
        iri = self.resolver.resolve(curie, "class", self.row_id).iri
        return Expr("named", iri=iri)


class Resolver:
    def __init__(self) -> None:
        self.graphs: dict[str, Graph] = {}
        self.label_index: dict[tuple[str, str, str], set[URIRef]] = defaultdict(set)
        self.labels: dict[URIRef, str] = {}
        self.records: dict[tuple[str, str, str], LabelResolutionRecord] = {}
        self.cache: dict[tuple[str, str], Resolution] = {}
        self.source_graph = Graph()
        self._load_graphs()

    def _load_graphs(self) -> None:
        loaded_by_file: dict[Path, Graph] = {}
        for prefix, rel_path in PREFIX_FILES.items():
            graph = loaded_by_file.get(rel_path)
            if graph is None:
                graph = Graph()
                graph.parse(REPO_ROOT / rel_path, format="turtle")
                loaded_by_file[rel_path] = graph
            self.graphs[prefix] = graph

        for path in SOURCE_IMPORTS:
            self.source_graph.parse(REPO_ROOT / path, format="turtle")

        for prefix, graph in self.graphs.items():
            if prefix in {"owl", "rdf", "rdfs"}:
                continue
            for subject, _, label in graph.triples((None, RDFS.label, None)):
                if not isinstance(subject, URIRef):
                    continue
                label_text = str(label)
                self.labels.setdefault(subject, label_text)
                for kind in self.kinds(subject, graph):
                    self.label_index[(prefix, self._normalize_key(label_text), kind)].add(subject)

    def _normalize_key(self, value: str) -> str:
        return re.sub(r"[^a-z0-9]+", "", value.lower())

    def _label_style_key(self, local: str) -> str:
        spaced = re.sub(r"(?<=[a-z0-9])(?=[A-Z])", " ", local)
        spaced = spaced.replace("_", " ").replace("-", " ")
        return self._normalize_key(spaced)

    def kinds(self, iri: URIRef, graph: Graph | None = None) -> set[str]:
        graph = graph or self.graph_for_iri(iri)
        kinds: set[str] = set()
        if (iri, RDF.type, OWL.Class) in graph or (iri, RDF.type, RDFS.Class) in graph:
            kinds.add("class")
        if (iri, RDF.type, OWL.ObjectProperty) in graph:
            kinds.add("object_property")
        return kinds

    def graph_for_iri(self, iri: URIRef) -> Graph:
        text = str(iri)
        for prefix, namespace in sorted(PREFIXES.items(), key=lambda item: len(item[1]), reverse=True):
            if prefix in self.graphs and text.startswith(namespace):
                return self.graphs[prefix]
        return self.graphs["cco"]

    def resolve(self, token: str, expected_kind: str, row_id: str) -> Resolution:
        cache_key = (token, expected_kind)
        if cache_key in self.cache:
            resolution = self.cache[cache_key]
            self._record_resolution(resolution, row_id)
            return resolution

        if ":" not in token:
            raise GenerationError(f"{row_id}: token {token!r} is not a CURIE")
        prefix, local = token.split(":", 1)
        if prefix not in PREFIXES:
            raise GenerationError(f"{row_id}: unknown prefix {prefix!r} in {token!r}")
        if prefix in {"owl", "rdf", "rdfs"}:
            iri = URIRef(PREFIXES[prefix] + local)
            resolution = Resolution(token, iri, expected_kind, "direct")
            self.cache[cache_key] = resolution
            return resolution
        graph = self.graphs.get(prefix)
        if graph is None:
            raise GenerationError(f"{row_id}: prefix {prefix!r} is not associated with a local ontology")

        direct_iri = URIRef(PREFIXES[prefix] + local)
        direct_kinds = self.kinds(direct_iri, graph)
        if expected_kind in direct_kinds:
            resolution = Resolution(token, direct_iri, expected_kind, "direct")
            self.cache[cache_key] = resolution
            return resolution
        if direct_kinds and expected_kind not in direct_kinds:
            raise GenerationError(
                f"{row_id}: {token} resolves directly to {sorted(direct_kinds)}, not {expected_kind}"
            )

        exact_matches = {
            subject
            for subject, label in graph.subject_objects(RDFS.label)
            if isinstance(subject, URIRef)
            and str(label) == local
            and expected_kind in self.kinds(subject, graph)
        }
        if len(exact_matches) == 1:
            iri = next(iter(exact_matches))
            resolution = Resolution(token, iri, expected_kind, "exact_label", self.labels.get(iri, local))
            self.cache[cache_key] = resolution
            self._record_resolution(resolution, row_id)
            return resolution
        if len(exact_matches) > 1:
            raise GenerationError(
                f"{row_id}: {token} has multiple exact label matches: "
                + ", ".join(sorted(map(str, exact_matches)))
            )

        label_key = self._label_style_key(local)
        style_matches = self.label_index.get((prefix, label_key, expected_kind), set())
        if len(style_matches) == 1:
            iri = next(iter(style_matches))
            resolution = Resolution(
                token,
                iri,
                expected_kind,
                "label_style_key",
                self.labels.get(iri, local),
            )
            self.cache[cache_key] = resolution
            self._record_resolution(resolution, row_id)
            return resolution
        if len(style_matches) > 1:
            raise GenerationError(
                f"{row_id}: {token} has multiple label-style matches: "
                + ", ".join(sorted(map(str, style_matches)))
            )

        raise GenerationError(f"{row_id}: unresolved {expected_kind} token {token!r}")

    def resolve_source_subject(self, token: str, row_id: str) -> tuple[URIRef, str]:
        if ":" not in token:
            raise GenerationError(f"{row_id}: source subject {token!r} is not a CURIE")
        prefix, _ = token.split(":", 1)
        if prefix not in {"sosa", "sampling", "ssn", "ssn-system"}:
            raise GenerationError(f"{row_id}: source subject {token!r} is not in a source ontology prefix")
        iri: URIRef | None = None
        source_kinds: set[str] = set()
        try:
            iri = self.resolve(token, "class", row_id).iri
            source_kinds = self.kinds(iri, self.source_graph)
        except GenerationError:
            pass
        if not source_kinds:
            try:
                iri = self.resolve(token, "object_property", row_id).iri
                source_kinds = self.kinds(iri, self.source_graph)
            except GenerationError:
                raise GenerationError(
                    f"{row_id}: source subject {token!r} cannot be resolved as a declared "
                    "OWL class or object property in the local SOSA/SSN source ontologies"
                ) from None
        if len(source_kinds) != 1:
            raise GenerationError(f"{row_id}: source subject {token!r} has ambiguous or missing type {source_kinds}")
        assert iri is not None
        return iri, next(iter(source_kinds))

    def _record_resolution(self, resolution: Resolution, row_id: str) -> None:
        if resolution.method == "direct":
            return
        key = (resolution.token, str(resolution.iri), resolution.kind)
        record = self.records.get(key)
        if record is None:
            record = LabelResolutionRecord(
                token=resolution.token,
                iri=resolution.iri,
                kind=resolution.kind,
                method=resolution.method,
                label=resolution.label,
            )
            self.records[key] = record
        record.rows.add(row_id)


def normalize_cell(value: object) -> str:
    if value is None:
        return ""
    return str(value).strip()


def compact_iri(value: URIRef | str) -> str:
    text = str(value)
    for prefix, namespace in sorted(PREFIXES.items(), key=lambda item: len(item[1]), reverse=True):
        if text.startswith(namespace):
            return f"{prefix}:{text[len(namespace):]}"
    return f"<{text}>"


def command_text(command: list[str]) -> str:
    return shlex.join(command)


def markdown_escape(value: str) -> str:
    return value.replace("|", "\\|").replace("\n", "<br>")


def sha256_file(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for block in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(block)
    return digest.hexdigest()


def bind_prefixes(graph: Graph) -> None:
    for prefix, namespace in PREFIXES.items():
        graph.bind(prefix, Namespace(namespace))


def read_workbook(path: Path) -> tuple[list[WorkbookRow], WorkbookStats]:
    workbook = openpyxl.load_workbook(path, data_only=True)
    rows: list[WorkbookRow] = []
    stats = WorkbookStats()
    for worksheet in workbook.worksheets:
        headers = [normalize_cell(worksheet.cell(row=1, column=col).value) for col in range(1, worksheet.max_column + 1)]
        if not all(column in headers for column in BASE_REQUIRED_COLUMNS):
            continue
        if ROW_ID_HEADER not in headers:
            workbook.close()
            raise GenerationError(
                f"{worksheet.title}!1: governed worksheet is missing required header {ROW_ID_HEADER!r}"
            )
        stats.worksheets_read.append(worksheet.title)
        header_index = {header: idx + 1 for idx, header in enumerate(headers)}
        for row_number in range(2, worksheet.max_row + 1):
            row = WorkbookRow(
                sheet=worksheet.title,
                row_number=row_number,
                subject_text=normalize_cell(worksheet.cell(row=row_number, column=header_index["sssom:subject_id"]).value),
                predicate_text=normalize_cell(worksheet.cell(row=row_number, column=header_index["sssom:predicate_id"]).value),
                target_text=normalize_cell(worksheet.cell(row=row_number, column=header_index["coms:Target"]).value),
                reasoning_text=normalize_cell(worksheet.cell(row=row_number, column=header_index["coms:Reasoning"]).value),
                stable_row_id=normalize_cell(worksheet.cell(row=row_number, column=header_index[ROW_ID_HEADER]).value),
                mapping_status_text=(
                    normalize_cell(
                        worksheet.cell(
                            row=row_number,
                            column=header_index["coms:MappingStatus"],
                        ).value
                    )
                    if "coms:MappingStatus" in header_index
                    else ""
                ),
            )
            stats.rows_by_sheet[worksheet.title] += 1
            if (
                row.subject_text
                or row.predicate_text
                or row.target_text
                or row.reasoning_text
                or row.mapping_status_text
                or row.stable_row_id
            ):
                stats.populated_rows_by_sheet[worksheet.title] += 1
                rows.append(row)
    workbook.close()
    if not stats.worksheets_read:
        raise GenerationError(f"no worksheet in {path} contains the required COMS header")
    return rows, stats


def validate_workbook_row_ids(rows: list[WorkbookRow], stats: WorkbookStats) -> None:
    references: list[RowIdentityReference] = []
    issues = []
    for row in rows:
        try:
            row_id = validate_row_id(row.stable_row_id, row.location)
        except ComsRowIdentityError as exc:
            issues.extend(exc.issues)
            continue
        references.append(RowIdentityReference(row_id=row_id, location=row.location))
    issues.extend(validate_unique_row_ids(references))
    stats.governed_row_id_count = len(rows)
    stats.unique_row_id_count = len({reference.row_id for reference in references})
    if issues:
        raise GenerationError(str(ComsRowIdentityError(issues)))


def validate_and_process_rows(rows: list[WorkbookRow], resolver: Resolver, stats: WorkbookStats) -> list[ProcessedRow]:
    validate_workbook_row_ids(rows, stats)
    processed: list[ProcessedRow] = []
    property_typing_row_by_key: dict[tuple[str, str], WorkbookRow] = {}
    for row in rows:
        if not row.subject_text:
            raise GenerationError(
                f"ERROR [MISSING_SOURCE_SUBJECT] {row.diagnostic_id}: "
                "sssom:subject_id is required for every governed row"
            )

        subject, subject_kind = resolver.resolve_source_subject(row.subject_text, row.diagnostic_id)

        if row.is_blank_mapping:
            stats.blank_mapping_rows += 1
            processed.append(
                ProcessedRow(
                    row=row,
                    subject=subject,
                    subject_kind=subject_kind,
                    predicate="",
                    target="",
                )
            )
            continue

        if not row.predicate_text or not row.target_text:
            raise GenerationError(
                f"{row.diagnostic_id}: active axiom rows must populate subject, predicate, and target; "
                "only subject-only rows are allowed as explicit blank mappings"
            )
        if row.predicate_text not in ALLOWED_PREDICATES:
            raise GenerationError(f"{row.diagnostic_id}: invalid predicate {row.predicate_text!r}")

        if row.predicate_text in CLASS_PREDICATES and subject_kind != "class":
            raise GenerationError(f"{row.diagnostic_id}: class predicate used with {subject_kind} subject")
        if row.predicate_text in OBJECT_PROPERTY_SUBJECT_PREDICATES and subject_kind != "object_property":
            raise GenerationError(
                f"{row.diagnostic_id}: {row.predicate_text} requires an object-property subject; "
                f"{row.subject_text} resolves as {subject_kind}"
            )

        key = (str(subject), row.predicate_text)
        if row.predicate_text in DOMAIN_RANGE_PREDICATES:
            previous_row = property_typing_row_by_key.get(key)
            if previous_row is not None:
                axiom_name = "domain" if row.predicate_text == "rdfs:domain" else "range"
                raise GenerationError(
                    f"{row.diagnostic_id}: duplicate {row.predicate_text} row for {row.subject_text}; "
                    f"the first {axiom_name} row is {previous_row.diagnostic_id}. Multiple OWL {axiom_name} "
                    "axioms are conjunctive; write alternatives with Manchester 'or' in one target expression."
                )
            property_typing_row_by_key[key] = row
        if row.predicate_text in CLASS_PREDICATES:
            expr = ManchesterParser(row.target_text, resolver, row.diagnostic_id).parse()
            stats.class_mapping_rows += 1
            stats.mapped_rows += 1
            processed.append(
                ProcessedRow(
                    row=row,
                    subject=subject,
                    subject_kind=subject_kind,
                    predicate=row.predicate_text,
                    target=row.target_text,
                    expr=expr,
                )
            )
        elif row.predicate_text in DOMAIN_RANGE_PREDICATES:
            expr = ManchesterParser(row.target_text, resolver, row.diagnostic_id).parse()
            if row.predicate_text == "rdfs:domain":
                stats.domain_rows += 1
            else:
                stats.range_rows += 1
            processed.append(
                ProcessedRow(
                    row=row,
                    subject=subject,
                    subject_kind=subject_kind,
                    predicate=row.predicate_text,
                    target=row.target_text,
                    expr=expr,
                )
            )
        elif row.predicate_text == "owl:propertyChainAxiom":
            chain = parse_property_chain(row.target_text, resolver, row.diagnostic_id)
            stats.property_chain_rows += 1
            stats.mapped_rows += 1
            processed.append(
                ProcessedRow(
                    row=row,
                    subject=subject,
                    subject_kind=subject_kind,
                    predicate=row.predicate_text,
                    target=row.target_text,
                    property_chain=chain,
                )
            )
        else:
            target_property = resolver.resolve(row.target_text, "object_property", row.diagnostic_id).iri
            stats.object_property_mapping_rows += 1
            stats.mapped_rows += 1
            processed.append(
                ProcessedRow(
                    row=row,
                    subject=subject,
                    subject_kind=subject_kind,
                    predicate=row.predicate_text,
                    target=row.target_text,
                    target_property=target_property,
                )
            )
    stats.processed_row_count = len(processed)
    identity_audits = attach_canonical_identities(processed)
    stats.identity_audit_row_count = len(identity_audits)
    validate_identity_audit_completeness(rows, processed, identity_audits)
    stats.identity_count_reconciliation_passed = True
    stats.identity_row_id_set_reconciliation_passed = True
    stats.identity_location_reconciliation_passed = True
    validate_incompatible_duplicate_mappings(processed)
    return processed


def canonical_expression_node(expr: Expr) -> CanonicalExpressionNode:
    if expr.kind == "named":
        return CanonicalExpressionNode(kind="named", iri=None if expr.iri is None else str(expr.iri))
    if expr.kind in {"intersection", "union"}:
        return CanonicalExpressionNode(
            kind=expr.kind,
            children=tuple(canonical_expression_node(child) for child in expr.children),
        )
    if expr.kind == "some":
        return CanonicalExpressionNode(
            kind="some",
            property_iri=None if expr.prop is None else str(expr.prop),
            filler=None if expr.filler is None else canonical_expression_node(expr.filler),
        )
    return CanonicalExpressionNode(kind=expr.kind)


def mapping_type_for_processed_row(item: ProcessedRow) -> str:
    if not item.predicate:
        return "explicit_blank"
    if item.predicate in CLASS_PREDICATES:
        return "class_mapping"
    if item.predicate == "owl:propertyChainAxiom":
        return "property_chain"
    if item.predicate == "rdfs:domain":
        return "domain"
    if item.predicate == "rdfs:range":
        return "range"
    return "object_property_mapping"


def canonical_input_for_processed_row(item: ProcessedRow) -> CanonicalRowInput:
    return CanonicalRowInput(
        row_id=item.row.stable_row_id,
        location=item.row.location,
        subject_iri=str(item.subject),
        predicate_iri=None if not item.predicate else str(ALLOWED_PREDICATES[item.predicate]),
        mapping_type=mapping_type_for_processed_row(item),
        reasoning=item.row.reasoning_text,
        expression=None if item.expr is None else canonical_expression_node(item.expr),
        target_property_iri=None if item.target_property is None else str(item.target_property),
        property_chain=tuple(str(value) for value in item.property_chain),
    )


def attach_canonical_identities(processed_rows: list[ProcessedRow]) -> tuple[CanonicalRowAudit, ...]:
    audits: list[CanonicalRowAudit] = []
    for item in processed_rows:
        try:
            audit = build_row_audit(canonical_input_for_processed_row(item))
        except ComsRowIdentityError as exc:
            raise GenerationError(str(exc)) from exc
        item.identity_audit = audit
        audits.append(audit)
    issues = validate_unique_authoritative_axioms(audits)
    if issues:
        raise GenerationError(str(ComsRowIdentityError(issues)))
    return tuple(audits)


def disposition_input_for_processed_row(item: ProcessedRow) -> DispositionRowInput:
    if item.identity_audit is None:
        raise GenerationError(f"{item.row.diagnostic_id}: identity audit is required for dispositions")
    canonical_input = canonical_input_for_processed_row(item)
    try:
        axiom_inputs = tuple(
            axiom_input_from_canonical_row(identity, canonical_input)
            for identity in item.identity_audit.authoritative_axioms
        )
    except ProductDispositionError as exc:
        raise GenerationError(str(exc)) from exc
    return DispositionRowInput(
        row_id=item.row.stable_row_id,
        location=item.row.location,
        subject_lexical=item.row.subject_text,
        predicate_lexical=item.row.predicate_text or None,
        authoritative_target_lexical=item.row.target_text or None,
        canonical_row=item.identity_audit.expression,
        source_expression_sha256=item.identity_audit.source_expression_sha256,
        mapping_type=mapping_type_for_processed_row(item),
        reasoning=item.row.reasoning_text,
        authoritative_axioms=axiom_inputs,
    )


def build_and_write_disposition_report(
    processed_rows: list[ProcessedRow],
    path: Path,
    input_hashes: RequiredInputHashes,
    publication_metadata: PublicationMetadata,
) -> tuple[DispositionDocument, list[DispositionRowInput]]:
    try:
        row_inputs = [disposition_input_for_processed_row(item) for item in processed_rows]
        document = build_disposition_document(
            row_inputs,
            publication_metadata,
            input_hashes,
        )
        path.parent.mkdir(parents=True, exist_ok=True)
        path.write_bytes(serialize_disposition_document(document))
        loaded, issues = validate_disposition_file(
            path,
            row_inputs,
            publication_metadata,
            input_hashes,
        )
    except (ProductDispositionError, PublicationMetadataError) as exc:
        raise GenerationError(str(exc)) from exc
    if issues:
        raise GenerationError(str(ProductDispositionError(issues)))
    return loaded, row_inputs


def validate_identity_audit_completeness(
    governed_rows: list[WorkbookRow],
    processed_rows: list[ProcessedRow],
    identity_audits: tuple[CanonicalRowAudit, ...],
) -> None:
    problems: list[tuple[str, str]] = []
    if len(governed_rows) != len(processed_rows):
        problems.append(
            (
                "IDENTITY_AUDIT_INCOMPLETE",
                f"governed row count {len(governed_rows)} does not match "
                f"processed row count {len(processed_rows)}",
            )
        )
    if len(processed_rows) != len(identity_audits):
        problems.append(
            (
                "IDENTITY_AUDIT_INCOMPLETE",
                f"processed row count {len(processed_rows)} does not match "
                f"identity-audit row count {len(identity_audits)}",
            )
        )

    governed_by_id = {
        row.stable_row_id: row.location
        for row in governed_rows
    }
    processed_by_id = {
        item.row.stable_row_id: item.row.location
        for item in processed_rows
    }
    audit_by_id = {
        audit.row_id: audit
        for audit in identity_audits
    }

    def located_ids(row_ids: set[str], locations: dict[str, RowLocation]) -> str:
        return ", ".join(
            f"{row_id} ({locations[row_id].text})"
            for row_id in sorted(row_ids)
        )

    governed_ids = set(governed_by_id)
    processed_ids = set(processed_by_id)
    audited_ids = set(audit_by_id)

    missing_processed = governed_ids - processed_ids
    unexpected_processed = processed_ids - governed_ids
    if missing_processed or unexpected_processed:
        details: list[str] = []
        if missing_processed:
            details.append(
                "governed RowIDs missing from processed rows: "
                + located_ids(missing_processed, governed_by_id)
            )
        if unexpected_processed:
            details.append(
                "unexpected processed RowIDs not present in governed rows: "
                + located_ids(unexpected_processed, processed_by_id)
            )
        problems.append(("GOVERNED_PROCESSED_ROWID_MISMATCH", "; ".join(details)))

    missing_audits = processed_ids - audited_ids
    unexpected_audits = audited_ids - processed_ids
    if missing_audits or unexpected_audits:
        details = []
        if missing_audits:
            details.append(
                "processed RowIDs missing from identity audits: "
                + located_ids(missing_audits, processed_by_id)
            )
        if unexpected_audits:
            audit_locations = {
                row_id: audit_by_id[row_id].location
                for row_id in unexpected_audits
            }
            details.append(
                "unexpected audited RowIDs not present in processed rows: "
                + located_ids(unexpected_audits, audit_locations)
            )
        problems.append(("PROCESSED_AUDIT_ROWID_MISMATCH", "; ".join(details)))

    for row_id in sorted(governed_ids & processed_ids):
        governed_location = governed_by_id[row_id]
        processed_location = processed_by_id[row_id]
        if governed_location != processed_location:
            problems.append(
                (
                    "IDENTITY_AUDIT_LOCATION_MISMATCH",
                    f"governed-to-processed location mismatch for {row_id}: "
                    f"governed {governed_location.text}, processed {processed_location.text}",
                )
            )

    for row_id in sorted(processed_ids & audited_ids):
        processed_location = processed_by_id[row_id]
        audit_location = audit_by_id[row_id].location
        if processed_location != audit_location:
            problems.append(
                (
                    "IDENTITY_AUDIT_LOCATION_MISMATCH",
                    f"processed-to-audit location mismatch for {row_id}: "
                    f"processed {processed_location.text}, audit {audit_location.text}",
                )
            )

    for item in processed_rows:
        audit = audit_by_id.get(item.row.stable_row_id)
        if audit is None or item.identity_audit is None:
            problems.append(
                (
                    "IDENTITY_AUDIT_INCOMPLETE",
                    f"{item.row.diagnostic_id} is absent from the identity audit",
                )
            )
            continue
        if item.identity_audit != audit:
            problems.append(
                (
                    "IDENTITY_AUDIT_INCOMPLETE",
                    f"{item.row.diagnostic_id} has a mismatched identity audit",
                )
            )
        if (
            audit.expression.mapping_type != "explicit_blank"
            and not audit.authoritative_axioms
        ):
            problems.append(
                (
                    "IDENTITY_AUDIT_INCOMPLETE",
                    f"{item.row.diagnostic_id} does not produce a canonical authoritative axiom",
                )
            )
        if (
            audit.expression.mapping_type == "explicit_blank"
            and audit.authoritative_axioms
        ):
            problems.append(
                (
                    "IDENTITY_AUDIT_INCOMPLETE",
                    f"{item.row.diagnostic_id} is an explicit blank row but "
                    "produces canonical authoritative axioms",
                )
            )

    if problems:
        raise GenerationError(
            " | ".join(f"ERROR [{code}] {message}" for code, message in problems)
        )


def validate_incompatible_duplicate_mappings(processed_rows: list[ProcessedRow]) -> None:
    first_by_key: dict[tuple[str, str], ProcessedRow] = {}
    for item in processed_rows:
        if not item.predicate or item.predicate in DOMAIN_RANGE_PREDICATES:
            continue
        key = (str(item.subject), item.predicate)
        previous = first_by_key.get(key)
        if previous is None:
            first_by_key[key] = item
            continue
        raise GenerationError(
            f"{item.row.diagnostic_id}: duplicate mapping for {item.row.subject_text} {item.predicate} "
            f"has an incompatible target; the first mapping row is {previous.row.diagnostic_id}. "
            "Canonically identical targets are reported as DUPLICATE_AUTHORITATIVE_AXIOM."
        )


def parse_property_chain(text: str, resolver: Resolver, row_id: str) -> tuple[URIRef, ...]:
    tokens = [token for token in re.split(r"\s+", text.strip()) if token]
    if len(tokens) < 3 or len(tokens) % 2 == 0:
        raise GenerationError(f"{row_id}: invalid property-chain syntax {text!r}")
    chain: list[URIRef] = []
    for index, token in enumerate(tokens):
        if index % 2 == 1:
            if token != "o":
                raise GenerationError(f"{row_id}: property-chain separator must be 'o', found {token!r}")
            continue
        chain.append(resolver.resolve(token, "object_property", row_id).iri)
    if len(chain) < 2:
        raise GenerationError(f"{row_id}: property chains must contain at least two properties")
    return tuple(chain)


def expr_to_rdf(graph: Graph, expr: Expr) -> URIRef | BNode:
    if expr.kind == "named":
        assert expr.iri is not None
        return expr.iri
    if expr.kind in {"intersection", "union"}:
        node = BNode()
        graph.add((node, RDF.type, OWL.Class))
        list_node = BNode()
        predicate = OWL.intersectionOf if expr.kind == "intersection" else OWL.unionOf
        graph.add((node, predicate, list_node))
        Collection(graph, list_node, [expr_to_rdf(graph, child) for child in expr.children])
        return node
    if expr.kind == "some":
        assert expr.prop is not None and expr.filler is not None
        node = BNode()
        graph.add((node, RDF.type, OWL.Restriction))
        graph.add((node, OWL.onProperty, expr.prop))
        graph.add((node, OWL.someValuesFrom, expr_to_rdf(graph, expr.filler)))
        return node
    raise GenerationError(f"unsupported expression node {expr.kind!r}")


ROOT_PREFIX_ORDER = (
    "bfo",
    "cco",
    "owl",
    "rdf",
    "rdfs",
    "sampling",
    "sosa",
    "ssn",
    "ssn-system",
)
ROOT_PREFIXES = (
    *METADATA_PREFIXES,
    *((prefix, PREFIXES[prefix]) for prefix in ROOT_PREFIX_ORDER),
)
ROOT_ORDERED_IMPORTS = tuple(sorted((str(value) for value in DIRECT_IMPORTS)))
ROOT_IMPORT_TURTLE_TERMS = (
    "sampling:",
    "ssn:",
    "ssn-system:",
    "<https://www.commoncoreontologies.org/2024-11-06/CommonCoreOntologiesMerged>",
)


def _root_turtle_bytes(
    processed_rows: Iterable[ProcessedRow],
    publication_metadata: PublicationMetadata,
    context: FormalReleaseContext | None = None,
) -> bytes:
    """Render maintained root bytes solely from governed structured inputs."""

    rendered_axioms: list[tuple[tuple[str, str, str], tuple[str, ...]]] = []
    for item in processed_rows:
        if not item.predicate:
            continue
        audit = item.identity_audit
        if audit is None or len(audit.authoritative_axioms) != 1:
            raise GenerationError(
                f"{item.row.diagnostic_id}: deterministic root rendering requires "
                "exactly one audited authoritative axiom"
            )
        canonical_input = canonical_input_for_processed_row(item)
        axiom_id = f"sha256:{audit.authoritative_axioms[0].sha256}"
        try:
            lines = render_authoritative_axiom_lines(canonical_input, axiom_id)
        except ModularProductError as exc:
            raise GenerationError(f"deterministic root rendering failed: {exc}") from exc
        rendered_axioms.append(
            (
                (
                    canonical_input.subject_iri,
                    canonical_input.predicate_iri or "",
                    axiom_id,
                ),
                lines,
            )
        )

    imports = (
        release_project_imports(publication_metadata, "integrated", context)
        if context
        else ROOT_ORDERED_IMPORTS
    )
    header = render_ontology_header_bytes(
        publication_metadata,
        "integrated",
        imports,
        generated_notice=GENERATED_NOTICE,
        prefixes=ROOT_PREFIXES,
        import_turtle_terms=ROOT_IMPORT_TURTLE_TERMS,
        context=context,
    )
    output = bytearray(header.rstrip(b"\n"))
    for _, lines in sorted(rendered_axioms, key=lambda value: value[0]):
        output.extend(b"\n\n")
        output.extend("\n".join(lines).encode("utf-8"))
    output.extend(b"\n")
    return bytes(output)


def render_formal_product_set(
    processed_rows: list[ProcessedRow],
    identity_audits: Iterable[CanonicalRowAudit],
    disposition_document: DispositionDocument,
    publication_metadata: PublicationMetadata,
    context: FormalReleaseContext,
) -> FormalProductSet:
    """Render all four formal products in memory from one governed input state."""

    validated_context = validate_formal_release_context(context)
    canonical_rows = tuple(
        canonical_input_for_processed_row(item) for item in processed_rows
    )
    audits = tuple(identity_audits)
    reconciliations = tuple(
        reconcile_product_axioms(
            product_key,
            canonical_rows,
            audits,
            disposition_document,
        )
        for product_key in (
            "alignment_core",
            "strict_bfo_mapping",
            "bfo_projection",
            "cco_extension",
        )
    )
    by_key = {value.product_key: value for value in reconciliations}

    root_bytes = _root_turtle_bytes(
        processed_rows,
        publication_metadata,
        validated_context,
    )
    root_graph = Graph().parse(data=root_bytes.decode("utf-8"), format="turtle")
    root_imports = release_project_imports(
        publication_metadata,
        "integrated",
        validated_context,
    )
    root_header_issues = validate_serialized_ontology_header(
        root_bytes,
        publication_metadata,
        "integrated",
        root_imports,
        generated_notice=GENERATED_NOTICE,
        prefixes=ROOT_PREFIXES,
        import_turtle_terms=ROOT_IMPORT_TURTLE_TERMS,
        mode="release",
        context=validated_context,
    )
    if root_header_issues:
        raise GenerationError(
            "formal integrated-root metadata validation failed: "
            + " | ".join(
                f"ERROR [{value.code}] {value.field}: {value.message}"
                for value in root_header_issues
            )
        )
    root_logical = strip_emitted_ontology_header(
        root_graph,
        publication_metadata,
        "integrated",
        root_imports,
        validated_context,
    )
    if len(root_logical) != ROOT_LOGICAL_TRIPLE_COUNT or len(root_graph) != ROOT_FORMAL_TOTAL_TRIPLE_COUNT:
        raise GenerationError(
            "formal integrated-root partition mismatch: "
            f"logical={len(root_logical)} total={len(root_graph)}"
        )

    core = build_alignment_core(
        by_key["alignment_core"].selected_axioms,
        publication_metadata,
        validated_context,
    )
    strict = build_strict_bfo_mapping(
        by_key["strict_bfo_mapping"].selected_axioms,
        publication_metadata,
        validated_context,
    )
    cco = build_cco_extension(
        by_key["cco_extension"].selected_axioms,
        publication_metadata,
        validated_context,
    )
    expected_totals = {
        "alignment_core": 64,
        "strict_bfo_mapping": 137,
        "cco_extension": 936,
    }
    for result in (core, strict, cco):
        expected_total = expected_totals[result.metadata.product_key]
        if result.total_triple_count != expected_total:
            raise GenerationError(
                f"formal {result.metadata.product_key} total triple count mismatch: "
                f"expected {expected_total}, got {result.total_triple_count}"
            )

    product_validation_issues = {
        "alignment_core": validate_alignment_core(
            core.serialized_bytes,
            by_key["alignment_core"].selected_axioms,
            publication_metadata,
            integrated_graph=root_graph,
            context=validated_context,
        ),
        "strict_bfo_mapping": validate_strict_bfo_mapping(
            strict.serialized_bytes,
            by_key["strict_bfo_mapping"].selected_axioms,
            core.serialized_bytes,
            by_key["alignment_core"].selected_axioms,
            publication_metadata,
            integrated_graph=root_graph,
            context=validated_context,
        ),
        "cco_extension": validate_cco_extension(
            cco.serialized_bytes,
            by_key["cco_extension"].selected_axioms,
            strict.serialized_bytes,
            by_key["strict_bfo_mapping"].selected_axioms,
            core.serialized_bytes,
            by_key["alignment_core"].selected_axioms,
            publication_metadata,
            integrated_graph=root_graph,
            context=validated_context,
        ),
    }
    for product_key, issues in product_validation_issues.items():
        if issues:
            raise GenerationError(
                f"formal {product_key} validation failed: "
                + " | ".join(
                    f"ERROR [{value.code}] {value.field}: {value.message}"
                    for value in issues
                )
            )

    integrated = RenderedOntologyProduct(
        product_key="integrated",
        serialized_bytes=root_bytes,
        ontology_declaration_triple_count=1,
        import_triple_count=4,
        metadata_annotation_count=7,
        formal_metadata_annotation_count=3,
        logical_triple_count=ROOT_LOGICAL_TRIPLE_COUNT,
        total_triple_count=len(root_graph),
        governed_axiom_count=103,
        sha256=hashlib.sha256(root_bytes).hexdigest(),
    )
    return FormalProductSet(
        context=validated_context,
        integrated=integrated,
        alignment_core=core,
        strict_bfo_mapping=strict,
        cco_extension=cco,
        reconciliations=reconciliations,
    )


def generate_ontology(
    processed_rows: list[ProcessedRow],
    output_path: Path,
    publication_metadata: PublicationMetadata,
    *,
    require_current_counts: bool = False,
) -> Graph:
    graph = Graph()
    bind_prefixes(graph)
    graph.add((ONTOLOGY_IRI, RDF.type, OWL.Ontology))
    for import_iri in DIRECT_IMPORTS:
        graph.add((ONTOLOGY_IRI, OWL.imports, import_iri))

    for item in processed_rows:
        if not item.predicate:
            continue
        predicate_iri = ALLOWED_PREDICATES[item.predicate]
        if item.expr is not None:
            graph.add((item.subject, predicate_iri, expr_to_rdf(graph, item.expr)))
        elif item.target_property is not None:
            graph.add((item.subject, predicate_iri, item.target_property))
        elif item.property_chain:
            chain_node = BNode()
            Collection(graph, chain_node, list(item.property_chain))
            graph.add((item.subject, OWL.propertyChainAxiom, chain_node))
        else:
            raise GenerationError(f"{item.row.diagnostic_id}: processed row has no target")

    logical_before_metadata = Graph()
    header_triples = {
        (ONTOLOGY_IRI, RDF.type, OWL.Ontology),
        *((ONTOLOGY_IRI, OWL.imports, value) for value in DIRECT_IMPORTS),
    }
    for triple in graph:
        if triple not in header_triples:
            logical_before_metadata.add(triple)

    serialized = _root_turtle_bytes(processed_rows, publication_metadata)
    parsed = Graph().parse(data=serialized.decode("utf-8"), format="turtle")
    expected_imports = ROOT_ORDERED_IMPORTS
    metadata_issues = validate_serialized_ontology_header(
        serialized,
        publication_metadata,
        "integrated",
        expected_imports,
        generated_notice=GENERATED_NOTICE,
        prefixes=ROOT_PREFIXES,
        import_turtle_terms=ROOT_IMPORT_TURTLE_TERMS,
    )
    if metadata_issues:
        raise GenerationError(
            "integrated-root metadata validation failed: "
            + " | ".join(
                f"ERROR [{value.code}] {value.field}: {value.message}"
                for value in metadata_issues
            )
        )
    logical_after_metadata = strip_emitted_ontology_header(
        parsed,
        publication_metadata,
        "integrated",
        expected_imports,
    )
    if require_current_counts and len(logical_after_metadata) != ROOT_LOGICAL_TRIPLE_COUNT:
        raise GenerationError(
            "integrated-root logical triple count mismatch: "
            f"expected {ROOT_LOGICAL_TRIPLE_COUNT}, got {len(logical_after_metadata)}"
        )
    if not isomorphic(logical_before_metadata, logical_after_metadata):
        raise GenerationError("integrated-root logical graph changed during metadata emission")
    if require_current_counts and len(parsed) != ROOT_TOTAL_TRIPLE_COUNT:
        raise GenerationError(
            f"integrated-root total triple count mismatch: expected {ROOT_TOTAL_TRIPLE_COUNT}, "
            f"got {len(parsed)}"
        )

    output_path.parent.mkdir(parents=True, exist_ok=True)
    output_path.write_bytes(serialized)
    return parsed


def normalized_axiom_rows(processed_rows: list[ProcessedRow], graph: Graph) -> list[NormalizedRow]:
    rows: list[NormalizedRow] = []
    for item in processed_rows:
        if not item.predicate:
            continue
        predicate_iri = ALLOWED_PREDICATES[item.predicate]
        if item.expr is not None:
            objects = sorted(
                (canonical_expr(graph, obj) for obj in graph.objects(item.subject, predicate_iri)),
            )
            normalized = " ; ".join(objects)
        elif item.target_property is not None:
            normalized = compact_iri(item.target_property)
        elif item.property_chain:
            normalized = " o ".join(compact_iri(prop) for prop in item.property_chain)
        else:
            normalized = ""
        rows.append(
            NormalizedRow(
                row_id=item.row.row_id,
                subject=compact_iri(item.subject),
                subject_kind=item.subject_kind,
                predicate=item.predicate,
                original_target=item.target,
                normalized_target=normalized,
                rdf_owl_form=f"{compact_iri(item.subject)} {item.predicate} {normalized} .",
            )
        )
    return rows


def unsat_classes(graph: Graph) -> list[URIRef]:
    classes: set[URIRef] = set()
    for subject in graph.subjects(RDFS.subClassOf, OWL.Nothing):
        if isinstance(subject, URIRef) and subject != OWL.Nothing:
            classes.add(subject)
    for subject in graph.subjects(OWL.equivalentClass, OWL.Nothing):
        if isinstance(subject, URIRef) and subject != OWL.Nothing:
            classes.add(subject)
    for obj in graph.objects(OWL.Nothing, OWL.equivalentClass):
        if isinstance(obj, URIRef) and obj != OWL.Nothing:
            classes.add(obj)
    return sorted(classes, key=str)


@dataclass(frozen=True)
class HermitRunProfile:
    graph_filename: str
    reasoned_filename: str


CANDIDATE_HERMIT_PROFILE = HermitRunProfile(
    graph_filename="coms-candidate-full-closure.ttl",
    reasoned_filename="coms-candidate-full-closure-reasoned.ttl",
)
ALIGNMENT_CORE_HERMIT_PROFILE = HermitRunProfile(
    graph_filename="alignment-core-source-closure.ttl",
    reasoned_filename="alignment-core-source-closure-reasoned.ttl",
)
STRICT_BFO_HERMIT_PROFILE = HermitRunProfile(
    graph_filename="strict-bfo-pinned-merged-cco-bfo-closure.ttl",
    reasoned_filename="strict-bfo-pinned-merged-cco-bfo-closure-reasoned.ttl",
)
CCO_EXTENSION_HERMIT_PROFILE = HermitRunProfile(
    graph_filename="cco-extension-pinned-merged-cco-bfo-closure.ttl",
    reasoned_filename="cco-extension-pinned-merged-cco-bfo-closure-reasoned.ttl",
)


def _run_hermit_closure(
    generated_path: Path,
    closure: Graph,
    tmp_dir: Path,
    profile: HermitRunProfile,
    *,
    validate_profile: bool = True,
) -> HermitResult:
    """Evaluate one exact closure with OWL 2 DL profile and HermiT gates."""

    generated_graph = Graph().parse(generated_path, format="turtle")

    tmp_dir.mkdir(parents=True, exist_ok=True)
    graph_path = tmp_dir / profile.graph_filename
    reasoned_path = tmp_dir / profile.reasoned_filename
    profile_graph_path = graph_path.with_name(
        f"{graph_path.stem}-owl2dl-profile{graph_path.suffix}"
    )

    reasoned_path.unlink(missing_ok=True)
    profile_graph_path.unlink(missing_ok=True)

    # This graph is the exact product-plus-pinned-dependency closure used by
    # HermiT and reported through closure_triple_count.
    closure.serialize(destination=graph_path, format="turtle")

    profile_triple_count: int | None = None
    profile_completion_count = 0
    profile_return_code: int | None = None
    profile_output = ""
    declarations_retained: bool | None = None

    if validate_profile:
        declarations_retained = all(
            triple in closure
            for triple in SAMPLE_PROPERTY_SOURCE_DECLARATIONS
        )

        profile_graph = Graph()
        bind_prefixes(profile_graph)
        for prefix, namespace in closure.namespaces():
            profile_graph.bind(prefix, namespace)
        for triple in closure:
            profile_graph.add(triple)

        for triple in DL_PROFILE_DECLARATION_COMPLETION:
            if triple not in profile_graph:
                profile_graph.add(triple)
                profile_completion_count += 1

        profile_triple_count = len(profile_graph)
        profile_graph.serialize(
            destination=profile_graph_path,
            format="turtle",
        )

    robot = shutil.which("robot")
    if robot is None:
        return HermitResult(
            graph_path=graph_path,
            reasoned_path=reasoned_path,
            generated_triple_count=len(generated_graph),
            closure_triple_count=len(closure),
            return_code=None,
            reasoned_output_produced=False,
            owl_nothing_count=None,
            unsat_classes=[],
            robot_output="ROBOT executable not found on PATH.",
            robot_path=None,
            profile_checked=validate_profile,
            profile_graph_path=profile_graph_path if validate_profile else None,
            profile_triple_count=profile_triple_count,
            profile_declaration_completion_count=profile_completion_count,
            profile_return_code=None,
            profile_output="ROBOT executable not found on PATH.",
            source_sample_declarations_retained=declarations_retained,
        )

    if validate_profile:
        if declarations_retained is not True:
            profile_return_code = 1
            profile_output = (
                "Exact validation closure does not retain both pinned SOSA "
                "sample-property declarations."
            )
        else:
            profile_proc = subprocess.run(
                [
                    robot,
                    "validate-profile",
                    "--profile",
                    "DL",
                    "--input",
                    str(profile_graph_path),
                ],
                cwd=REPO_ROOT,
                text=True,
                stdout=subprocess.PIPE,
                stderr=subprocess.PIPE,
            )
            profile_return_code = profile_proc.returncode
            profile_output = "\n".join(
                part
                for part in (
                    profile_proc.stdout.strip(),
                    profile_proc.stderr.strip(),
                )
                if part
            )

        if profile_return_code != 0:
            return HermitResult(
                graph_path=graph_path,
                reasoned_path=reasoned_path,
                generated_triple_count=len(generated_graph),
                closure_triple_count=len(closure),
                return_code=None,
                reasoned_output_produced=False,
                owl_nothing_count=None,
                unsat_classes=[],
                robot_output=profile_output,
                robot_path=robot,
                profile_checked=True,
                profile_graph_path=profile_graph_path,
                profile_triple_count=profile_triple_count,
                profile_declaration_completion_count=profile_completion_count,
                profile_return_code=profile_return_code,
                profile_output=profile_output,
                source_sample_declarations_retained=declarations_retained,
            )

    proc = subprocess.run(
        [
            robot,
            "reason",
            "--reasoner",
            "HermiT",
            "--input",
            str(graph_path),
            "--output",
            str(reasoned_path),
        ],
        cwd=REPO_ROOT,
        text=True,
        stdout=subprocess.PIPE,
        stderr=subprocess.PIPE,
    )

    reasoner_output = "\n".join(
        part
        for part in (
            proc.stdout.strip(),
            proc.stderr.strip(),
        )
        if part
    )
    combined_output = "\n".join(
        part
        for part in (
            profile_output.strip(),
            reasoner_output.strip(),
        )
        if part
    )

    inferred_unsats = {
        URIRef(match.group(1))
        for match in UNSAT_RE.finditer(reasoner_output)
    }

    reasoned_output_produced = (
        reasoned_path.exists()
        and reasoned_path.stat().st_size > 0
    )
    owl_nothing_count: int | None = None

    if reasoned_output_produced:
        reasoned_graph = Graph()
        bind_prefixes(reasoned_graph)
        reasoned_graph.parse(reasoned_path, format="turtle")
        inferred_unsats |= set(unsat_classes(reasoned_graph))
        owl_nothing_count = len(inferred_unsats)

    return HermitResult(
        graph_path=graph_path,
        reasoned_path=reasoned_path,
        generated_triple_count=len(generated_graph),
        closure_triple_count=len(closure),
        return_code=proc.returncode,
        reasoned_output_produced=reasoned_output_produced,
        owl_nothing_count=owl_nothing_count,
        unsat_classes=sorted(inferred_unsats, key=str),
        robot_output=combined_output,
        robot_path=robot,
        profile_checked=validate_profile,
        profile_graph_path=profile_graph_path if validate_profile else None,
        profile_triple_count=profile_triple_count,
        profile_declaration_completion_count=profile_completion_count,
        profile_return_code=profile_return_code,
        profile_output=profile_output,
        source_sample_declarations_retained=declarations_retained,
    )

def run_candidate_hermit(
    generated_path: Path,
    tmp_dir: Path,
) -> HermitResult:
    closure = Graph()
    bind_prefixes(closure)

    for input_path in CANDIDATE_CLOSURE_INPUTS:
        closure.parse(
            REPO_ROOT / input_path,
            format="turtle",
        )

    closure.parse(generated_path, format="turtle")

    for triple in list(
        closure.triples((None, OWL.imports, None))
    ):
        closure.remove(triple)

    return _run_hermit_closure(
        generated_path,
        closure,
        tmp_dir,
        CANDIDATE_HERMIT_PROFILE,
    )


def run_alignment_core_hermit(
    generated_path: Path,
    tmp_dir: Path,
) -> HermitResult:
    """Reason over the import-free core plus the fixed tracked source closure."""

    closure = build_fixed_source_closure(
        generated_path.read_bytes(),
        (REPO_ROOT / path for path in SOURCE_IMPORTS),
    )

    return _run_hermit_closure(
        generated_path,
        closure,
        tmp_dir,
        ALIGNMENT_CORE_HERMIT_PROFILE,
    )


def run_strict_bfo_hermit(
    generated_path: Path,
    alignment_core_path: Path,
    tmp_dir: Path,
) -> HermitResult:
    """Reason over the strict product and its explicit pinned validation closure."""

    closure = build_fixed_validation_closure(
        (
            generated_path.read_bytes(),
            alignment_core_path.read_bytes(),
        ),
        (
            REPO_ROOT / BFO_VALIDATION_DEPENDENCY,
            *(REPO_ROOT / path for path in SOURCE_IMPORTS),
        ),
    )

    return _run_hermit_closure(
        generated_path,
        closure,
        tmp_dir,
        STRICT_BFO_HERMIT_PROFILE,
    )


def run_cco_extension_hermit(
    generated_path: Path,
    strict_bfo_path: Path,
    alignment_core_path: Path,
    tmp_dir: Path,
) -> HermitResult:
    """Reason over the CCO extension and its explicit pinned validation closure."""

    closure = build_fixed_validation_closure(
        (
            generated_path.read_bytes(),
            strict_bfo_path.read_bytes(),
            alignment_core_path.read_bytes(),
        ),
        (
            REPO_ROOT / BFO_VALIDATION_DEPENDENCY,
            *(REPO_ROOT / path for path in SOURCE_IMPORTS),
        ),
    )

    return _run_hermit_closure(
        generated_path,
        closure,
        tmp_dir,
        CCO_EXTENSION_HERMIT_PROFILE,
    )




def build_and_write_alignment_core(
    processed_rows: list[ProcessedRow],
    identity_audits: list[CanonicalRowAudit],
    disposition_document: DispositionDocument,
    publication_metadata: PublicationMetadata,
    integrated_graph: Graph,
    output_path: Path,
    tmp_dir: Path,
) -> tuple[ModularProductResult, HermitResult]:
    """Build and fully validate the candidate alignment-core development artifact."""

    try:
        canonical_rows = [canonical_input_for_processed_row(item) for item in processed_rows]
        selected = select_product_axioms(
            "alignment_core",
            canonical_rows,
            identity_audits,
            disposition_document,
        )
        result = build_alignment_core(selected, publication_metadata)
        closure = build_fixed_source_closure(
            result.serialized_bytes,
            (REPO_ROOT / path for path in SOURCE_IMPORTS),
        )
        structural_issues = validate_alignment_core(
            result.serialized_bytes,
            selected,
            publication_metadata,
            fixed_source_closure=closure,
            integrated_graph=integrated_graph,
        )
        if structural_issues:
            raise ModularProductError(structural_issues)
        output_path.parent.mkdir(parents=True, exist_ok=True)
        output_path.write_bytes(serialize_modular_product(result))
        hermit = run_alignment_core_hermit(output_path, tmp_dir)
    except (ModularProductError, PublicationMetadataError) as exc:
        output_path.unlink(missing_ok=True)
        raise GenerationError(str(exc)) from exc
    if not hermit.passed:
        output_path.unlink(missing_ok=True)
        raise GenerationError("alignment-core fixed source closure is not HermiT-clean")
    return result, hermit


def build_and_write_strict_bfo_mapping(
    processed_rows: list[ProcessedRow],
    identity_audits: list[CanonicalRowAudit],
    disposition_document: DispositionDocument,
    publication_metadata: PublicationMetadata,
    integrated_graph: Graph,
    alignment_core_result: ModularProductResult,
    alignment_core_path: Path,
    output_path: Path,
    tmp_dir: Path,
) -> tuple[ModularProductResult, HermitResult]:
    """Build and validate the strict BFO development artifact and fixed closure."""

    try:
        canonical_rows = [canonical_input_for_processed_row(item) for item in processed_rows]
        selected = select_product_axioms(
            "strict_bfo_mapping",
            canonical_rows,
            identity_audits,
            disposition_document,
        )
        alignment_selected = tuple(
            axiom
            for row in alignment_core_result.selected_rows
            for axiom in row.axioms
        )
        result = build_strict_bfo_mapping(selected, publication_metadata)
        fixed_closure = build_fixed_validation_closure(
            (result.serialized_bytes, alignment_core_result.serialized_bytes),
            (
                REPO_ROOT / BFO_VALIDATION_DEPENDENCY,
                *(REPO_ROOT / path for path in SOURCE_IMPORTS),
            ),
            )
        structural_issues = validate_strict_bfo_mapping(
            result.serialized_bytes,
            selected,
            alignment_core_result.serialized_bytes,
            alignment_selected,
            publication_metadata,
            integrated_graph=integrated_graph,
            fixed_semantic_closure=fixed_closure,
        )
        if structural_issues:
            raise ModularProductError(structural_issues)
        output_path.parent.mkdir(parents=True, exist_ok=True)
        output_path.write_bytes(serialize_modular_product(result))
        hermit = run_strict_bfo_hermit(output_path, alignment_core_path, tmp_dir)
    except (ModularProductError, PublicationMetadataError) as exc:
        output_path.unlink(missing_ok=True)
        raise GenerationError(str(exc)) from exc
    if not hermit.passed:
        output_path.unlink(missing_ok=True)
        raise GenerationError(
            "strict BFO mapping pinned merged CCO/BFO validation closure is not HermiT-clean"
        )
    return result, hermit




def build_and_write_cco_extension(
    processed_rows: list[ProcessedRow],
    identity_audits: list[CanonicalRowAudit],
    disposition_document: DispositionDocument,
    publication_metadata: PublicationMetadata,
    integrated_graph: Graph,
    alignment_core_result: ModularProductResult,
    alignment_core_path: Path,
    strict_bfo_result: ModularProductResult,
    strict_bfo_path: Path,
    output_path: Path,
    tmp_dir: Path,
) -> tuple[ModularProductResult, HermitResult]:
    """Build and validate the CCO-extension development artifact and closure."""

    try:
        canonical_rows = [canonical_input_for_processed_row(item) for item in processed_rows]
        selected = select_product_axioms(
            "cco_extension",
            canonical_rows,
            identity_audits,
            disposition_document,
        )
        alignment_selected = tuple(
            axiom
            for row in alignment_core_result.selected_rows
            for axiom in row.axioms
        )
        strict_selected = tuple(
            axiom
            for row in strict_bfo_result.selected_rows
            for axiom in row.axioms
        )
        result = build_cco_extension(selected, publication_metadata)
        source_graph = Graph()
        for path in SOURCE_IMPORTS:
            source_graph.parse(REPO_ROOT / path, format="turtle")
        merged_dependency = Graph().parse(
            REPO_ROOT / BFO_VALIDATION_DEPENDENCY,
            format="turtle",
        )
        fixed_closure = build_fixed_validation_closure(
            (
                result.serialized_bytes,
                strict_bfo_result.serialized_bytes,
                alignment_core_result.serialized_bytes,
            ),
            (
                REPO_ROOT / BFO_VALIDATION_DEPENDENCY,
                *(REPO_ROOT / path for path in SOURCE_IMPORTS),
            ),
            )
        structural_issues = validate_cco_extension(
            result.serialized_bytes,
            selected,
            strict_bfo_result.serialized_bytes,
            strict_selected,
            alignment_core_result.serialized_bytes,
            alignment_selected,
            publication_metadata,
            integrated_graph=integrated_graph,
            fixed_semantic_closure=fixed_closure,
            source_dependency_graph=source_graph,
            merged_cco_bfo_dependency_graph=merged_dependency,
        )
        if structural_issues:
            raise ModularProductError(structural_issues)
        output_path.parent.mkdir(parents=True, exist_ok=True)
        output_path.write_bytes(serialize_modular_product(result))
        hermit = run_cco_extension_hermit(
            output_path,
            strict_bfo_path,
            alignment_core_path,
            tmp_dir,
        )
    except (ModularProductError, PublicationMetadataError) as exc:
        output_path.unlink(missing_ok=True)
        raise GenerationError(str(exc)) from exc
    if not hermit.passed:
        output_path.unlink(missing_ok=True)
        raise GenerationError(
            "CCO extension pinned merged CCO/BFO validation closure is not HermiT-clean"
        )
    return result, hermit


def run_select_query(graph: Graph, query_path: Path) -> list[dict[str, str]]:
    rows: list[dict[str, str]] = []
    for row in graph.query((REPO_ROOT / query_path).read_text(encoding="utf-8")):
        asdict = row.asdict()
        rows.append({key: str(value) for key, value in asdict.items()})
    return rows


def build_source_graph() -> Graph:
    graph = Graph()
    bind_prefixes(graph)
    for path in SOURCE_IMPORTS:
        graph.parse(REPO_ROOT / path, format="turtle")
    return graph


def build_coverage(
    processed_rows: list[ProcessedRow],
    source_subject_errors: Iterable[str],
    coverage_report_path: Path,
) -> CoverageResult:
    source_graph = build_source_graph()
    source_query_rows = run_select_query(source_graph, Path("queries/source-classes-and-object-properties.rq"))
    source_terms = {
        URIRef(row["term"]): row["kind"]
        for row in source_query_rows
    }

    mapped_terms = {row.subject for row in processed_rows if row.predicate in MAPPING_PREDICATES}
    property_typing_terms = {
        row.subject
        for row in processed_rows
        if row.predicate in DOMAIN_RANGE_PREDICATES
    }
    listed_terms = {row.subject for row in processed_rows}
    explicit_blank_terms = {row.subject for row in processed_rows if not row.predicate}
    missing_subjects = set(source_subject_errors)

    coverage_graph = Graph()
    bind_prefixes(coverage_graph)
    coverage_graph.bind("coms", COMS_COVERAGE)
    for term, kind in sorted(source_terms.items(), key=lambda item: str(item[0])):
        coverage_graph.add((term, COMS_COVERAGE.sourceKind, Literal(kind)))
        if term in mapped_terms:
            status = "mapped"
        elif term in property_typing_terms:
            status = "covered_by_property_typing"
        elif term in explicit_blank_terms:
            status = "explicitly_unmapped"
        else:
            status = "absent_from_spreadsheet"
        coverage_graph.add((term, COMS_COVERAGE.coverageStatus, Literal(status)))

    unmapped_rows = run_select_query(coverage_graph, Path("queries/unmapped-source-terms.rq"))
    result = CoverageResult(
        source_terms=source_terms,
        mapped_terms=mapped_terms,
        property_typing_terms=property_typing_terms,
        listed_terms=listed_terms,
        explicit_blank_terms=explicit_blank_terms,
        spreadsheet_missing_subjects=missing_subjects,
        query_source_count=len(source_query_rows),
        query_unmapped_count=len(unmapped_rows),
    )
    write_coverage_report(coverage_report_path, result)
    return result


def format_term_list(values: Iterable[URIRef | str]) -> list[str]:
    return [f"- `{compact_iri(value)}`" for value in sorted(values, key=str)]


def write_coverage_report(path: Path, coverage: CoverageResult) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    lines = [
        "# COMS Source-Term Coverage",
        "",
        "This report is generated by `tools/generate_mapping_from_coms.py`.",
        "",
        "Coverage scope is all non-deprecated named OWL classes and object properties defined by:",
        "",
        *[f"- `{path}`" for path in SOURCE_IMPORTS],
        "",
        "The source term inventory is produced by `queries/source-classes-and-object-properties.rq`. "
        "Unmapped terms are selected from the generated coverage graph by `queries/unmapped-source-terms.rq`.",
        "A domain or range row covers a property for source-term coverage but does not count it as relation-mapped; only subproperty, equivalent-property, or property-chain rows do so.",
        "",
        "## Summary",
        "",
        "| Item | Count |",
        "|---|---:|",
        f"| source terms returned by SPARQL | {coverage.query_source_count} |",
        f"| mapped classes | {len(coverage.mapped_classes)} |",
        f"| unmapped classes | {len(coverage.unmapped_classes)} |",
        f"| mapped object properties | {len(coverage.mapped_object_properties)} |",
        f"| unmapped object properties | {len(coverage.unmapped_object_properties)} |",
        f"| explicitly listed blank mappings | {len(coverage.explicit_blank_terms)} |",
        f"| listed only in domain/range property-typing rows | {len(coverage.property_typing_only_terms)} |",
        f"| source terms absent from spreadsheet | {len(coverage.absent_terms)} |",
        f"| spreadsheet subjects not found in source ontologies | {len(coverage.spreadsheet_missing_subjects)} |",
        f"| unmapped rows returned by SPARQL coverage query | {coverage.query_unmapped_count} |",
        "",
        "## Mapped Classes",
        "",
        *format_term_list(coverage.mapped_classes),
        "",
        "## Unmapped Classes",
        "",
        *format_term_list(coverage.unmapped_classes),
        "",
        "## Mapped Object Properties",
        "",
        *format_term_list(coverage.mapped_object_properties),
        "",
        "## Unmapped Object Properties",
        "",
        *format_term_list(coverage.unmapped_object_properties),
        "",
        "## Explicitly Listed Blank Mappings",
        "",
        *format_term_list(coverage.explicit_blank_terms),
        "",
        "## Listed Only In Domain/Range Property-Typing Rows",
        "",
        *format_term_list(coverage.property_typing_only_terms),
        "",
        "## Source Terms Absent From Spreadsheet",
        "",
        *format_term_list(coverage.absent_terms),
        "",
        "## Spreadsheet Subjects Not Found In Source Ontologies",
        "",
        *(f"- `{value}`" for value in sorted(coverage.spreadsheet_missing_subjects)),
        "",
    ]
    path.write_text("\n".join(lines), encoding="utf-8")


def canonical_expr(graph: Graph, node: URIRef | BNode) -> str:
    if isinstance(node, URIRef):
        return str(node)
    if (node, OWL.intersectionOf, None) in graph:
        list_node = next(graph.objects(node, OWL.intersectionOf))
        items = sorted(canonical_expr(graph, item) for item in Collection(graph, list_node))
        return "ObjectIntersectionOf(" + " ".join(items) + ")"
    if (node, OWL.unionOf, None) in graph:
        list_node = next(graph.objects(node, OWL.unionOf))
        items = sorted(canonical_expr(graph, item) for item in Collection(graph, list_node))
        return "ObjectUnionOf(" + " ".join(items) + ")"
    if (node, OWL.onProperty, None) in graph and (node, OWL.someValuesFrom, None) in graph:
        prop = next(graph.objects(node, OWL.onProperty))
        filler = next(graph.objects(node, OWL.someValuesFrom))
        return f"ObjectSomeValuesFrom({prop} {canonical_expr(graph, filler)})"
    return f"_:{node}"


def property_chain_text(graph: Graph, list_node: BNode) -> str:
    return " o ".join(str(item) for item in Collection(graph, list_node))


def extract_mapping_axioms(graph: Graph) -> set[tuple[str, str, str, str]]:
    mappings: set[tuple[str, str, str, str]] = set()
    for predicate in (RDFS.subClassOf, OWL.equivalentClass):
        for subject, _, obj in graph.triples((None, predicate, None)):
            if isinstance(subject, URIRef):
                mappings.add(("class", str(subject), str(predicate), canonical_expr(graph, obj)))
    for predicate in (RDFS.subPropertyOf, OWL.equivalentProperty):
        for subject, _, obj in graph.triples((None, predicate, None)):
            if isinstance(subject, URIRef) and isinstance(obj, URIRef):
                mappings.add(("object_property", str(subject), str(predicate), str(obj)))
    for subject, _, obj in graph.triples((None, OWL.propertyChainAxiom, None)):
        if isinstance(subject, URIRef) and isinstance(obj, BNode):
            mappings.add(("property_chain", str(subject), str(OWL.propertyChainAxiom), property_chain_text(graph, obj)))
    return mappings


def extract_property_typing_axioms(graph: Graph, predicate: URIRef) -> set[tuple[str, str]]:
    return {
        (str(subject), canonical_expr(graph, obj))
        for subject, _, obj in graph.triples((None, predicate, None))
        if isinstance(subject, URIRef) and isinstance(obj, (URIRef, BNode))
    }


def compare_coms_to_legacy(
    generated_path: Path,
    report_path: Path,
    processed_rows: list[ProcessedRow],
) -> ComparisonResult:
    generated = Graph()
    generated.parse(generated_path, format="turtle")
    legacy = Graph()
    legacy.parse(LEGACY_ONTOLOGY, format="turtle")

    generated_mappings = extract_mapping_axioms(generated)
    legacy_mappings = extract_mapping_axioms(legacy)
    both = generated_mappings & legacy_mappings
    coms_only = generated_mappings - legacy_mappings
    legacy_only = legacy_mappings - generated_mappings

    class_diffs = diff_by_subject_predicate(generated_mappings, legacy_mappings, "class")
    object_property_diffs = diff_by_subject_predicate(generated_mappings, legacy_mappings, "object_property")
    property_chain_diffs = diff_by_subject_predicate(generated_mappings, legacy_mappings, "property_chain")

    generated_domains = extract_property_typing_axioms(generated, RDFS.domain)
    legacy_domains = extract_property_typing_axioms(legacy, RDFS.domain)
    generated_ranges = extract_property_typing_axioms(generated, RDFS.range)
    legacy_ranges = extract_property_typing_axioms(legacy, RDFS.range)

    result = ComparisonResult(
        both=both,
        coms_only=coms_only,
        legacy_only=legacy_only,
        class_expression_differences=class_diffs,
        object_property_differences=object_property_diffs,
        property_chain_differences=property_chain_diffs,
        domain_both=generated_domains & legacy_domains,
        domain_coms_only=generated_domains - legacy_domains,
        domain_legacy_only=legacy_domains - generated_domains,
        domain_differences=diff_axiom_targets(generated_domains, legacy_domains),
        range_both=generated_ranges & legacy_ranges,
        range_coms_only=generated_ranges - legacy_ranges,
        range_legacy_only=legacy_ranges - generated_ranges,
        range_differences=diff_axiom_targets(generated_ranges, legacy_ranges),
    )
    write_comparison_report(report_path, result, processed_rows)
    return result


def diff_by_subject_predicate(
    generated: set[tuple[str, str, str, str]],
    legacy: set[tuple[str, str, str, str]],
    kind: str,
) -> list[str]:
    gen_index = defaultdict(set)
    legacy_index = defaultdict(set)
    for item in generated:
        if item[0] == kind:
            gen_index[(item[1], item[2])].add(item[3])
    for item in legacy:
        if item[0] == kind:
            legacy_index[(item[1], item[2])].add(item[3])
    diffs: list[str] = []
    for key in sorted(set(gen_index) & set(legacy_index)):
        if gen_index[key] != legacy_index[key]:
            subject, predicate = key
            diffs.append(
                f"`{compact_iri(subject)}` `{compact_iri(predicate)}`: "
                f"COMS={sorted(gen_index[key])}; legacy={sorted(legacy_index[key])}"
            )
    return diffs


def diff_axiom_targets(
    generated: set[tuple[str, str]],
    legacy: set[tuple[str, str]],
) -> list[str]:
    gen_index: dict[str, set[str]] = defaultdict(set)
    legacy_index: dict[str, set[str]] = defaultdict(set)
    for subject, target in generated:
        gen_index[subject].add(target)
    for subject, target in legacy:
        legacy_index[subject].add(target)
    return [
        f"`{compact_iri(subject)}`: COMS={sorted(gen_index[subject])}; "
        f"legacy={sorted(legacy_index[subject])}"
        for subject in sorted(set(gen_index) & set(legacy_index))
        if gen_index[subject] != legacy_index[subject]
    ]


def format_mapping_rows(rows: Iterable[tuple[str, str, str, str]], limit: int | None = None) -> list[str]:
    formatted: list[str] = []
    sorted_rows = sorted(rows)
    if limit is not None:
        sorted_rows = sorted_rows[:limit]
    for kind, subject, predicate, target in sorted_rows:
        formatted.append(
            f"- `{kind}` `{compact_iri(subject)}` `{compact_iri(predicate)}` `{target}`"
        )
    return formatted or ["- none"]


def format_property_typing_rows(rows: Iterable[tuple[str, str]], predicate: str) -> list[str]:
    return [
        f"- `{compact_iri(subject)}` `{predicate}` `{target}`"
        for subject, target in sorted(rows)
    ] or ["- none"]


def write_comparison_report(path: Path, result: ComparisonResult, processed_rows: list[ProcessedRow]) -> None:
    blank_rows = [row for row in processed_rows if not row.predicate]
    class_difference_lines = [f"- {value}" for value in result.class_expression_differences] or ["- none"]
    object_property_difference_lines = [f"- {value}" for value in result.object_property_differences] or ["- none"]
    property_chain_difference_lines = [f"- {value}" for value in result.property_chain_differences] or ["- none"]
    domain_difference_lines = [f"- {value}" for value in result.domain_differences] or ["- none"]
    range_difference_lines = [f"- {value}" for value in result.range_differences] or ["- none"]
    blank_row_lines = [f"- `{item.row.subject_text}` at `{item.row.row_id}`" for item in blank_rows] or ["- none"]
    path.parent.mkdir(parents=True, exist_ok=True)
    lines = [
        "# COMS vs Pre-COMS Legacy Mapping Diff",
        "",
        "This informational report compares mapping-bearing axioms and, separately, domain/range property-typing axioms generated from `mappings/SSN2BFO-COMS.xlsx` for `SSN2BFO.ttl` against the frozen `legacy/SSN2BFO-pre-COMS.ttl` snapshot. "
        "COMS is not required to reproduce every legacy axiom, and the two ontologies are never loaded together for candidate validation.",
        "",
        "## Summary",
        "",
        "| Item | Count |",
        "|---|---:|",
        f"| mappings present in both | {len(result.both)} |",
        f"| mappings only in COMS | {len(result.coms_only)} |",
        f"| mappings only in pre-COMS legacy ontology | {len(result.legacy_only)} |",
        f"| class-expression differences | {len(result.class_expression_differences)} |",
        f"| object-property mapping differences | {len(result.object_property_differences)} |",
        f"| property-chain differences | {len(result.property_chain_differences)} |",
        f"| domain axioms present in both | {len(result.domain_both)} |",
        f"| domain axioms only in COMS | {len(result.domain_coms_only)} |",
        f"| domain axioms only in pre-COMS legacy ontology | {len(result.domain_legacy_only)} |",
        f"| domain target differences | {len(result.domain_differences)} |",
        f"| range axioms present in both | {len(result.range_both)} |",
        f"| range axioms only in COMS | {len(result.range_coms_only)} |",
        f"| range axioms only in pre-COMS legacy ontology | {len(result.range_legacy_only)} |",
        f"| range target differences | {len(result.range_differences)} |",
        f"| legacy domain/range axioms absent from COMS | {len(result.legacy_domain_range_absent_from_coms)} |",
        f"| spreadsheet rows intentionally producing no mapping | {len(blank_rows)} |",
        "",
        "## Mappings Present In Both",
        "",
        *format_mapping_rows(result.both),
        "",
        "## Only In COMS",
        "",
        *format_mapping_rows(result.coms_only),
        "",
        "## Only In Pre-COMS Legacy Ontology",
        "",
        *format_mapping_rows(result.legacy_only),
        "",
        "## Class-Expression Differences",
        "",
        *class_difference_lines,
        "",
        "## Object-Property Mapping Differences",
        "",
        *object_property_difference_lines,
        "",
        "## Property-Chain Differences",
        "",
        *property_chain_difference_lines,
        "",
        "## Domain Axioms Present In Both",
        "",
        *format_property_typing_rows(result.domain_both, "rdfs:domain"),
        "",
        "## Domain Axioms Only In COMS",
        "",
        *format_property_typing_rows(result.domain_coms_only, "rdfs:domain"),
        "",
        "## Domain Axioms Only In Pre-COMS Legacy Ontology",
        "",
        *format_property_typing_rows(result.domain_legacy_only, "rdfs:domain"),
        "",
        "## Domain Target Differences",
        "",
        *domain_difference_lines,
        "",
        "## Range Axioms Present In Both",
        "",
        *format_property_typing_rows(result.range_both, "rdfs:range"),
        "",
        "## Range Axioms Only In COMS",
        "",
        *format_property_typing_rows(result.range_coms_only, "rdfs:range"),
        "",
        "## Range Axioms Only In Pre-COMS Legacy Ontology",
        "",
        *format_property_typing_rows(result.range_legacy_only, "rdfs:range"),
        "",
        "## Range Target Differences",
        "",
        *range_difference_lines,
        "",
        "## Spreadsheet Rows Intentionally Producing No Mapping",
        "",
        *blank_row_lines,
        "",
        "## Terms Requiring Human Review",
        "",
        "Human review should consider COMS-versus-legacy mapping differences separately from domain/range property-typing differences. Legacy-only axioms are informational and are not release requirements.",
        "",
    ]
    path.write_text("\n".join(lines), encoding="utf-8")


def write_generation_report(
    path: Path,
    *,
    workbook_path: Path,
    stats: WorkbookStats,
    resolver: Resolver,
    errors: list[str],
    output_path: Path,
    hermit: HermitResult | None,
    coverage: CoverageResult | None,
    comparison: ComparisonResult | None,
    normalized_rows: list[NormalizedRow],
    identity_audits: list[CanonicalRowAudit],
    elapsed_seconds: float,
    workbook_sha256: str,
    generator_sha256: str,
    identity_module_sha256: str,
    generation_timestamp: str,
    candidate_sha256: str,
    disposition_document: DispositionDocument | None = None,
    disposition_path: Path | None = None,
    disposition_sha256: str = "unavailable",
    disposition_module_sha256: str = "unavailable",
    publication_metadata_sha256: str = "unavailable",
    modular_products_module_sha256: str = "unavailable",
    alignment_core_result: ModularProductResult | None = None,
    alignment_core_path: Path | None = None,
    alignment_core_sha256: str = "unavailable",
    alignment_core_hermit: HermitResult | None = None,
    strict_bfo_result: ModularProductResult | None = None,
    strict_bfo_path: Path | None = None,
    strict_bfo_sha256: str = "unavailable",
    strict_bfo_hermit: HermitResult | None = None,
    cco_extension_result: ModularProductResult | None = None,
    cco_extension_path: Path | None = None,
    cco_extension_sha256: str = "unavailable",
    cco_extension_hermit: HermitResult | None = None,
) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    error_lines = [f"- {error}" for error in errors] or ["- none"]
    authoritative_axioms = [
        axiom.canonical_axiom
        for audit in identity_audits
        for axiom in audit.authoritative_axioms
    ]
    unique_authoritative_axioms = set(authoritative_axioms)
    lines = [
        "# COMS Mapping Generation Validation",
        "",
        "This report is generated by `tools/generate_mapping_from_coms.py`.",
        "",
        "## Source Metadata",
        "",
        "Freshness is determined from content hashes, not file timestamps.",
        "",
        "| Item | Value |",
        "|---|---|",
        f"| workbook SHA-256 | `{workbook_sha256}` |",
        f"| generator SHA-256 | `{generator_sha256}` |",
        f"| row-identity module SHA-256 | `{identity_module_sha256}` |",
        f"| product-disposition module SHA-256 | `{disposition_module_sha256}` |",
        f"| modular-products module SHA-256 | `{modular_products_module_sha256}` |",
        f"| publication metadata SHA-256 | `{publication_metadata_sha256}` |",
        f"| generation timestamp (UTC) | `{generation_timestamp}` |",
        f"| maintained ontology path | `{output_path}` |",
        f"| generated ontology SHA-256 | `{candidate_sha256}` |",
        f"| maintained product-disposition path | `{disposition_path or 'unavailable'}` |",
        f"| product-disposition JSON SHA-256 | `{disposition_sha256}` |",
        f"| maintained alignment-core path | `{alignment_core_path or 'unavailable'}` |",
        f"| alignment-core Turtle SHA-256 | `{alignment_core_sha256}` |",
        f"| maintained strict-BFO path | `{strict_bfo_path or 'unavailable'}` |",
        f"| strict-BFO Turtle SHA-256 | `{strict_bfo_sha256}` |",
        f"| maintained CCO-extension path | `{cco_extension_path or 'unavailable'}` |",
        f"| CCO-extension Turtle SHA-256 | `{cco_extension_sha256}` |",
        "",
        "## Workbook",
        "",
        f"- Workbook path: `{workbook_path}`",
        f"- Worksheets read: {', '.join(f'`{name}`' for name in stats.worksheets_read) or 'none'}",
        f"- Governed RowID header: `{ROW_ID_HEADER}` (recognized on {', '.join(f'`{name}`' for name in stats.worksheets_read) or 'no worksheets'})",
        "",
        "| Worksheet | Rows scanned | Populated rows |",
        "|---|---:|---:|",
        *[
            f"| `{sheet}` | {stats.rows_by_sheet[sheet]} | {stats.populated_rows_by_sheet[sheet]} |"
            for sheet in stats.worksheets_read
        ],
        "",
        "## Row Counts",
        "",
        "| Item | Count |",
        "|---|---:|",
        f"| active axiom row count | {stats.active_axiom_rows} |",
        f"| mapped row count | {stats.mapped_rows} |",
        f"| blank mapping row count | {stats.blank_mapping_rows} |",
        f"| class mapping count | {stats.class_mapping_rows} |",
        f"| object-property mapping count | {stats.object_property_mapping_rows} |",
        f"| domain row count | {stats.domain_rows} |",
        f"| range row count | {stats.range_rows} |",
        f"| property-chain count | {stats.property_chain_rows} |",
        f"| governed RowID count | {stats.governed_row_id_count} |",
        f"| unique RowID count | {stats.unique_row_id_count} |",
        f"| processed governed row count | {stats.processed_row_count} |",
        f"| identity-audit row count | {stats.identity_audit_row_count} |",
        "",
        "## Prefixes Derived",
        "",
        *[f"- `{prefix}:` `{namespace}`" for prefix, namespace in sorted(PREFIXES.items())],
        "",
        "## Label-To-IRI Resolutions",
        "",
    ]
    records = sorted(resolver.records.values(), key=lambda record: (record.token, str(record.iri), record.kind))
    if records:
        lines.extend(["| Token | Kind | Method | Label | IRI | Rows |", "|---|---|---|---|---|---|"])
        for record in records:
            lines.append(
                "| "
                + " | ".join(
                    [
                        f"`{record.token}`",
                        record.kind,
                        record.method,
                        record.label.replace("|", "\\|"),
                        f"`{record.iri}`",
                        ", ".join(f"`{row}`" for row in sorted(record.rows)),
                    ]
                )
                + " |"
            )
    else:
        lines.append("No label-to-IRI fallback resolutions were required.")

    if disposition_document is None:
        disposition_lines = [
            "## Product Dispositions",
            "",
            "- Disposition generation: unavailable",
        ]
    else:
        disposition_summary = disposition_document.summary
        disposition_lines = [
            "## Product Dispositions",
            "",
            f"- Path: `{disposition_path}`",
            "- Schema version: `1`",
            f"- Product order: {', '.join(f'`{key}`' for key in disposition_document.product_order)}",
            f"- Governed rows: {disposition_summary.governed_row_count}",
            f"- Canonical authoritative axioms: {disposition_summary.authoritative_axiom_count}",
            f"- Zero-axiom rows: {disposition_summary.zero_axiom_row_count}",
            f"- Target-neutral axioms: {disposition_summary.target_neutral_axiom_count}",
            f"- BFO-bearing axioms: {disposition_summary.bfo_bearing_axiom_count}",
            f"- CCO-bearing axioms: {disposition_summary.cco_bearing_axiom_count}",
            f"- Mixed BFO/CCO axioms: {disposition_summary.mixed_bfo_cco_axiom_count}",
            "- Disposition reconciliation and canonical serialization: PASS",
            "- No lossless transformation or weakened projection is approved by this artifact.",
        ]

    if alignment_core_result is None:
        alignment_core_lines = [
            "## Alignment Core",
            "",
            "- Generation and validation: unavailable",
        ]
    else:
        source_paths = ", ".join(f"`{path}`" for path in SOURCE_IMPORTS)
        alignment_core_lines = [
            "## Alignment Core",
            "",
            "This is the maintained authoritative development artifact at the approved production path; it is not a frozen formal release.",
            "",
            f"- Path: `{alignment_core_path}`",
            f"- Stable ontology IRI: `{alignment_core_result.metadata.stable_ontology_iri}`",
            f"- Governed authoritative axioms: {alignment_core_result.governed_axiom_count}",
            f"- Domain axioms: {alignment_core_result.domain_axiom_count}",
            f"- Range axioms: {alignment_core_result.range_axiom_count}",
            f"- Named target expressions: {alignment_core_result.named_target_count}",
            f"- Union target expressions: {alignment_core_result.union_target_count}",
            f"- Logical RDF triples: {alignment_core_result.logical_triple_count}",
            f"- Ontology declaration triples: {alignment_core_result.ontology_declaration_triple_count}",
            f"- Import triples: {alignment_core_result.import_triple_count}",
            f"- Descriptive metadata annotations: {alignment_core_result.metadata_annotation_count}",
            f"- Total RDF triples: {alignment_core_result.total_triple_count}",
            "- BFO/CCO/RO and unexpected logical-vocabulary audit: PASS",
            "- Integrated-root canonical-axiom reconciliation: PASS",
            "- Deterministic serialization: PASS",
            f"- Fixed local source closure: {source_paths}",
            f"- Source-closure triple count: {'n/a' if alignment_core_hermit is None else alignment_core_hermit.closure_triple_count}",
            f"- Source-closure HermiT return code: {'n/a' if alignment_core_hermit is None else alignment_core_hermit.return_code}",
            f"- Source-closure reasoned output produced: {'no' if alignment_core_hermit is None else 'yes' if alignment_core_hermit.reasoned_output_produced else 'no'}",
            f"- Source-closure named unsatisfiable classes: {'n/a' if alignment_core_hermit is None else len(alignment_core_hermit.unsat_classes)}",
            f"- Source-closure HermiT result: {'FAIL' if alignment_core_hermit is None else 'PASS' if alignment_core_hermit.passed else 'FAIL'}",
            "- Formal-release metadata and identity remain deferred.",
        ]

    if strict_bfo_result is None:
        strict_bfo_lines = [
            "## Strict BFO Mapping",
            "",
            "- Generation and validation: unavailable",
        ]
    else:
        closure_paths = ", ".join(
            f"`{path}`"
            for path in (*SOURCE_IMPORTS, BFO_VALIDATION_DEPENDENCY)
        )
        strict_bfo_lines = [
            "## Strict BFO Mapping",
            "",
            "This is the maintained authoritative development artifact at the approved production path; it is not a frozen formal release.",
            "",
            f"- Path: `{strict_bfo_path}`",
            f"- Stable ontology IRI: `{strict_bfo_result.metadata.stable_ontology_iri}`",
            f"- Direct governed authoritative axioms: {strict_bfo_result.governed_axiom_count}",
            f"- Subclass axioms: {strict_bfo_result.subclass_axiom_count}",
            f"- Equivalent-class axioms: {strict_bfo_result.equivalent_class_axiom_count}",
            f"- Direct subproperty axioms: {strict_bfo_result.direct_subproperty_axiom_count}",
            f"- Property-chain axioms: {strict_bfo_result.property_chain_axiom_count}",
            f"- Domain axioms: {strict_bfo_result.domain_axiom_count}",
            f"- Range axioms: {strict_bfo_result.range_axiom_count}",
            f"- Logical RDF triples: {strict_bfo_result.logical_triple_count}",
            f"- Ontology declaration triples: {strict_bfo_result.ontology_declaration_triple_count}",
            f"- Import triples: {strict_bfo_result.import_triple_count}",
            f"- Descriptive metadata annotations: {strict_bfo_result.metadata_annotation_count}",
            f"- Total RDF triples: {strict_bfo_result.total_triple_count}",
            "- Import: `http://www.sks.ai/SSN2BFO/current-ssn-sosa/alignment-core`",
            "- Imported alignment-core governed axioms: 29",
            "- Project-module closure governed axioms: 48",
            "- Direct/alignment-core governed overlap: 0",
            "- CCO/RO and unexpected logical-vocabulary audit: PASS",
            "- Integrated-root and project-module canonical-axiom reconciliation: PASS",
            "- Deterministic serialization: PASS",
            f"- Pinned merged CCO/BFO validation closure: {closure_paths}",
            f"- Pinned closure triple count: {'n/a' if strict_bfo_hermit is None else strict_bfo_hermit.closure_triple_count}",
            f"- HermiT return code: {'n/a' if strict_bfo_hermit is None else strict_bfo_hermit.return_code}",
            f"- Reasoned output produced: {'no' if strict_bfo_hermit is None else 'yes' if strict_bfo_hermit.reasoned_output_produced else 'no'}",
            f"- Named unsatisfiable classes: {'n/a' if strict_bfo_hermit is None else len(strict_bfo_hermit.unsat_classes)}",
            f"- HermiT result: {'FAIL' if strict_bfo_hermit is None else 'PASS' if strict_bfo_hermit.passed else 'FAIL'}",
            "- The validation dependency `imports/cco.ttl` is not imported by the published strict graph and does not authorize CCO mappings or transformations.",
            "- The 55 CCO-bearing or mixed axioms remain deferred; no transformation or projection is implemented.",
            "- Formal-release metadata and identity remain deferred.",
        ]

    if cco_extension_result is None:
        cco_extension_lines = [
            "## CCO Extension",
            "",
            "- Generation and validation: unavailable",
        ]
    else:
        closure_paths = ", ".join(
            f"`{path}`"
            for path in (*SOURCE_IMPORTS, BFO_VALIDATION_DEPENDENCY)
        )
        cco_selected = [
            axiom
            for row in cco_extension_result.selected_rows
            for axiom in row.axioms
        ]
        cco_bearing_count = sum(
            axiom.target_category == "cco_bearing" for axiom in cco_selected
        )
        mixed_count = sum(
            axiom.target_category == "mixed_bfo_cco" for axiom in cco_selected
        )
        cco_extension_lines = [
            "## CCO Extension",
            "",
            "This is the maintained authoritative development artifact at the approved production path; it is not a frozen formal release.",
            "",
            f"- Path: `{cco_extension_path}`",
            f"- Stable ontology IRI: `{cco_extension_result.metadata.stable_ontology_iri}`",
            f"- Direct governed authoritative axioms: {cco_extension_result.governed_axiom_count}",
            f"- CCO-bearing axioms: {cco_bearing_count}",
            f"- Mixed BFO/CCO axioms: {mixed_count}",
            f"- Subclass axioms: {cco_extension_result.subclass_axiom_count}",
            f"- Equivalent-class axioms: {cco_extension_result.equivalent_class_axiom_count}",
            f"- Direct subproperty axioms: {cco_extension_result.direct_subproperty_axiom_count}",
            f"- Property-chain axioms: {cco_extension_result.property_chain_axiom_count}",
            f"- Logical RDF triples: {cco_extension_result.logical_triple_count}",
            f"- Ontology declaration triples: {cco_extension_result.ontology_declaration_triple_count}",
            f"- Import triples: {cco_extension_result.import_triple_count}",
            f"- Descriptive metadata annotations: {cco_extension_result.metadata_annotation_count}",
            f"- Total RDF triples: {cco_extension_result.total_triple_count}",
            "- Import: `http://www.sks.ai/SSN2BFO/current-ssn-sosa/bfo-mapping`",
            "- Imported strict-BFO governed axioms: 19",
            "- Transitively imported alignment-core governed axioms: 29",
            "- Project-module closure governed axioms: 103",
            "- Pairwise governed overlap: 0",
            "- RO and unexpected logical-vocabulary audit: PASS",
            "- Integrated-root and project-module canonical-axiom reconciliation: PASS",
            "- Deterministic serialization: PASS",
            f"- Pinned merged CCO/BFO validation closure: {closure_paths}",
            f"- Pinned closure triple count: {'n/a' if cco_extension_hermit is None else cco_extension_hermit.closure_triple_count}",
            f"- HermiT return code: {'n/a' if cco_extension_hermit is None else cco_extension_hermit.return_code}",
            f"- Reasoned output produced: {'no' if cco_extension_hermit is None else 'yes' if cco_extension_hermit.reasoned_output_produced else 'no'}",
            f"- Named unsatisfiable classes: {'n/a' if cco_extension_hermit is None else len(cco_extension_hermit.unsat_classes)}",
            f"- HermiT result: {'FAIL' if cco_extension_hermit is None else 'PASS' if cco_extension_hermit.passed else 'FAIL'}",
            "- The validation dependency `imports/cco.ttl` is not imported by the published CCO-extension graph and is not mapping authority.",
            "- No transformation or weakened projection is implemented.",
            "- Formal-release metadata and identity remain deferred.",
        ]

    lines.extend(
        [
            "",
            "## Malformed Or Unresolved Rows",
            "",
            *error_lines,
            "",
            "## Duplicate-Row Checks",
            "",
            "- Duplicate subject/predicate rows with incompatible targets are fatal.",
            "- At most one populated `rdfs:domain` and one populated `rdfs:range` row are allowed per subject; alternatives belong in one Manchester `or` expression because multiple OWL domain/range axioms are conjunctive.",
            "- No incompatible duplicate mappings were found." if not errors else "- Duplicate checks did not complete cleanly because generation errors were present.",
            "",
            "## COMS Row Identity Integrity",
            "",
            f"- Canonical-expression version: `{CANONICALIZATION_VERSION}`",
            f"- Governed RowID count: {stats.governed_row_id_count}",
            f"- Unique RowID count: {stats.unique_row_id_count}",
            f"- Processed governed row count: {stats.processed_row_count}",
            f"- Identity-audit row count: {stats.identity_audit_row_count}",
            f"- Canonical authoritative axiom count: {len(authoritative_axioms)}",
            f"- Unique authoritative axiom count: {len(unique_authoritative_axioms)}",
            f"- Count reconciliation result: {'PASS' if stats.identity_count_reconciliation_passed else 'FAIL'}",
            f"- RowID-set reconciliation result: {'PASS' if stats.identity_row_id_set_reconciliation_passed else 'FAIL'}",
            f"- Location reconciliation result: {'PASS' if stats.identity_location_reconciliation_passed else 'FAIL'}",
            "- Identity-audit completeness result: "
            + (
                "PASS"
                if stats.identity_count_reconciliation_passed
                and stats.identity_row_id_set_reconciliation_passed
                and stats.identity_location_reconciliation_passed
                and all(audit.authoritative_axioms for audit in identity_audits)
                else "FAIL"
            ),
            f"- Duplicate RowID result: {'PASS' if stats.governed_row_id_count == stats.unique_row_id_count else 'FAIL'}",
            f"- Duplicate authoritative-axiom result: {'PASS' if len(identity_audits) == stats.governed_row_id_count and len(authoritative_axioms) == len(unique_authoritative_axioms) else 'FAIL'}",
            "",
            "The RowID is persistent identity; the source-expression SHA-256 excludes location and `coms:Reasoning`. Location and reasoning remain report-visible governed metadata.",
            "",
            "| RowID | Worksheet | Row | Mapping type | Source-expression SHA-256 | Canonical authoritative axiom | Reasoning |",
            "|---|---|---:|---|---|---|---|",
            *[
                "| "
                + " | ".join(
                    [
                        f"`{audit.row_id}`",
                        f"`{audit.location.worksheet}`",
                        str(audit.location.row_number),
                        f"`{audit.expression.mapping_type}`",
                        f"`{audit.source_expression_sha256}`",
                        "<br>".join(
                            "`" + markdown_escape(axiom.canonical_axiom) + "`"
                            for axiom in audit.authoritative_axioms
                        )
                        or "_(none)_",
                        markdown_escape(audit.reasoning) or "_(blank)_",
                    ]
                )
                + " |"
                for audit in sorted(
                    identity_audits,
                    key=lambda item: (item.location, item.row_id),
                )
            ],
            "",
            *disposition_lines,
            "",
            *alignment_core_lines,
            "",
            *strict_bfo_lines,
            "",
            "",
            *cco_extension_lines,
            "",
            "## Generated Ontology",
            "",
            f"- Path: `{output_path}`",
            f"- Ontology declaration triples: {ROOT_ONTOLOGY_DECLARATION_COUNT}",
            f"- Import triples: {ROOT_IMPORT_COUNT}",
            f"- Descriptive metadata annotations: {ROOT_METADATA_ANNOTATION_COUNT}",
            f"- Governed/structural logical triples: {ROOT_LOGICAL_TRIPLE_COUNT}",
            f"- Generated ontology triple count: {'n/a' if hermit is None else hermit.generated_triple_count}",
            f"- `{output_path}` is generated from `mappings/SSN2BFO-COMS.xlsx` and must not be edited directly.",
            "- `coms:Reasoning` remained spreadsheet-only and was not emitted into the ontology.",
            "",
            "## Candidate Closure HermiT Result",
            "",
            "| Item | Result |",
            "|---|---|",
            f"| ROBOT command | {robot_command_report_value(hermit)} |",
        ]
    )
    if hermit is None:
        lines.extend(
            [
                "| full candidate closure triple count | n/a |",
                "| HermiT return code | n/a |",
                "| reasoned output produced | no |",
                "| `owl:Nothing` count | n/a |",
                "| named unsat count | n/a |",
                "| named unsat set | n/a |",
                "| overall status | FAIL |",
            ]
        )
    else:
        lines.extend(
            [
                "| candidate closure graph path | temporary validation artifact (`coms-candidate-full-closure.ttl`) |",
                "| reasoned output path | temporary validation artifact (`coms-candidate-full-closure-reasoned.ttl`) |",
                f"| full candidate closure triple count | {hermit.closure_triple_count} |",
                f"| HermiT return code | {'' if hermit.return_code is None else hermit.return_code} |",
                f"| reasoned output produced | {'yes' if hermit.reasoned_output_produced else 'no'} |",
                f"| `owl:Nothing` count | {'n/a' if hermit.owl_nothing_count is None else hermit.owl_nothing_count} |",
                f"| named unsat count | {len(hermit.unsat_classes)} |",
                f"| named unsat set | {', '.join(f'`{compact_iri(value)}`' for value in hermit.unsat_classes) or 'clean'} |",
                f"| overall status | {'PASS' if hermit.passed else 'FAIL'} |",
            ]
        )
        if hermit.robot_output:
            excerpt = hermit.robot_output.strip()
            if len(excerpt) > 4000:
                excerpt = excerpt[:4000] + "\n... [truncated]"
            lines.extend(
                [
                    "",
                    "### ROBOT Output",
                    "",
                    "```text",
                    excerpt,
                    "```",
                ]
            )

    lines.extend(
        [
            "",
            "## Parsed/Normalized Axiom Expressions",
            "",
            "This section records mapping and domain/range property-typing rows after parsing, making Manchester grouping, subject typing, and generated RDF/OWL form visible during review.",
            "",
        ]
    )
    if normalized_rows:
        lines.extend(
            [
                "| Row | Subject | Predicate | Object-property subject? | Original target | Normalized target | Generated RDF/OWL form |",
                "|---|---|---|---|---|---|---|",
            ]
        )
        for row in normalized_rows:
            lines.append(
                "| "
                + " | ".join(
                    [
                        f"`{row.row_id}`",
                        f"`{row.subject}`",
                        f"`{row.predicate}`",
                        "yes" if row.subject_kind == "object_property" else "no",
                        row.original_target.replace("|", "\\|"),
                        row.normalized_target.replace("|", "\\|"),
                        "`" + row.rdf_owl_form.replace("|", "\\|") + "`",
                    ]
                )
                + " |"
            )
    else:
        lines.append("No active axiom rows were normalized.")

    lines.extend(
        [
            "",
            "## Source-Term Coverage Summary",
            "",
            "| Item | Count |",
            "|---|---:|",
        ]
    )
    if coverage is None:
        lines.extend(
            [
                "| mapped classes | n/a |",
                "| unmapped classes | n/a |",
                "| mapped object properties | n/a |",
                "| unmapped object properties | n/a |",
            ]
        )
    else:
        lines.extend(
            [
                f"| mapped classes | {len(coverage.mapped_classes)} |",
                f"| unmapped classes | {len(coverage.unmapped_classes)} |",
                f"| mapped object properties | {len(coverage.mapped_object_properties)} |",
                f"| unmapped object properties | {len(coverage.unmapped_object_properties)} |",
                f"| explicitly listed blank mappings | {len(coverage.explicit_blank_terms)} |",
                f"| listed only in domain/range property-typing rows | {len(coverage.property_typing_only_terms)} |",
                f"| source terms absent from spreadsheet | {len(coverage.absent_terms)} |",
                f"| spreadsheet subjects not found in source ontologies | {len(coverage.spreadsheet_missing_subjects)} |",
            ]
        )

    lines.extend(
        [
            "",
            "## COMS-Versus-Pre-COMS-Legacy Summary",
            "",
            "| Item | Count |",
            "|---|---:|",
        ]
    )
    if comparison is None:
        lines.extend(
            [
                "| mappings present in both | n/a |",
                "| mappings only in COMS | n/a |",
                "| mappings only in pre-COMS legacy ontology | n/a |",
            ]
        )
    else:
        lines.extend(
            [
                f"| mappings present in both | {len(comparison.both)} |",
                f"| mappings only in COMS | {len(comparison.coms_only)} |",
                f"| mappings only in pre-COMS legacy ontology | {len(comparison.legacy_only)} |",
                f"| class-expression differences | {len(comparison.class_expression_differences)} |",
                f"| object-property mapping differences | {len(comparison.object_property_differences)} |",
                f"| property-chain differences | {len(comparison.property_chain_differences)} |",
                f"| domain axioms present in both | {len(comparison.domain_both)} |",
                f"| domain axioms only in COMS | {len(comparison.domain_coms_only)} |",
                f"| domain axioms only in pre-COMS legacy ontology | {len(comparison.domain_legacy_only)} |",
                f"| domain target differences | {len(comparison.domain_differences)} |",
                f"| range axioms present in both | {len(comparison.range_both)} |",
                f"| range axioms only in COMS | {len(comparison.range_coms_only)} |",
                f"| range axioms only in pre-COMS legacy ontology | {len(comparison.range_legacy_only)} |",
                f"| range target differences | {len(comparison.range_differences)} |",
                f"| legacy domain/range axioms absent from COMS | {len(comparison.legacy_domain_range_absent_from_coms)} |",
            ]
        )

    lines.extend(
        [
            "",
            "## Runtime",
            "",
            f"- Runtime seconds: {elapsed_seconds:.2f}",
            "",
        ]
    )
    path.write_text("\n".join(lines), encoding="utf-8")


def write_summary_json(
    path: Path,
    *,
    workbook_path: Path,
    output_path: Path,
    workbook_sha256: str,
    generator_sha256: str,
    identity_module_sha256: str,
    generation_timestamp: str,
    candidate_sha256: str,
    stats: WorkbookStats,
    identity_audits: list[CanonicalRowAudit],
    hermit: HermitResult | None,
    coverage: CoverageResult | None,
    comparison: ComparisonResult | None,
    errors: list[str],
    elapsed_seconds: float,
    disposition_document: DispositionDocument | None = None,
    disposition_path: Path | None = None,
    disposition_sha256: str = "unavailable",
    disposition_module_sha256: str = "unavailable",
    publication_metadata_sha256: str = "unavailable",
    modular_products_module_sha256: str = "unavailable",
    alignment_core_result: ModularProductResult | None = None,
    alignment_core_path: Path | None = None,
    alignment_core_sha256: str = "unavailable",
    alignment_core_hermit: HermitResult | None = None,
    strict_bfo_result: ModularProductResult | None = None,
    strict_bfo_path: Path | None = None,
    strict_bfo_sha256: str = "unavailable",
    strict_bfo_hermit: HermitResult | None = None,
    cco_extension_result: ModularProductResult | None = None,
    cco_extension_path: Path | None = None,
    cco_extension_sha256: str = "unavailable",
    cco_extension_hermit: HermitResult | None = None,
) -> None:
    payload = {
        "status": "PASS" if not errors else "FAIL",
        "workbook_path": str(workbook_path),
        "output_path": str(output_path),
        "workbook_sha256": workbook_sha256,
        "generator_sha256": generator_sha256,
        "row_identity_module_sha256": identity_module_sha256,
        "product_disposition_module_sha256": disposition_module_sha256,
        "modular_products_module_sha256": modular_products_module_sha256,
        "publication_metadata_sha256": publication_metadata_sha256,
        "generation_timestamp": generation_timestamp,
        "generated_candidate_sha256": candidate_sha256,
        "generated_candidate_ontology_declaration_triple_count": ROOT_ONTOLOGY_DECLARATION_COUNT,
        "generated_candidate_import_triple_count": ROOT_IMPORT_COUNT,
        "generated_candidate_metadata_annotation_count": ROOT_METADATA_ANNOTATION_COUNT,
        "generated_candidate_logical_triple_count": ROOT_LOGICAL_TRIPLE_COUNT,
        "product_disposition_report_path": None if disposition_path is None else str(disposition_path),
        "product_disposition_report_sha256": disposition_sha256,
        "alignment_core_path": None if alignment_core_path is None else str(alignment_core_path),
        "alignment_core_sha256": alignment_core_sha256,
        "alignment_core": None
        if alignment_core_result is None
        else {
            "product_key": alignment_core_result.metadata.product_key,
            "stable_ontology_iri": alignment_core_result.metadata.stable_ontology_iri,
            "governed_axiom_count": alignment_core_result.governed_axiom_count,
            "logical_triple_count": alignment_core_result.logical_triple_count,
            "ontology_declaration_triple_count": alignment_core_result.ontology_declaration_triple_count,
            "import_triple_count": alignment_core_result.import_triple_count,
            "metadata_annotation_count": alignment_core_result.metadata_annotation_count,
            "total_triple_count": alignment_core_result.total_triple_count,
            "domain_axiom_count": alignment_core_result.domain_axiom_count,
            "range_axiom_count": alignment_core_result.range_axiom_count,
            "named_target_count": alignment_core_result.named_target_count,
            "union_target_count": alignment_core_result.union_target_count,
            "hermit_return_code": None if alignment_core_hermit is None else alignment_core_hermit.return_code,
            "hermit_result": "FAIL" if alignment_core_hermit is None else "PASS" if alignment_core_hermit.passed else "FAIL",
            "closure_triple_count": None if alignment_core_hermit is None else alignment_core_hermit.closure_triple_count,
            "named_unsat_count": None if alignment_core_hermit is None else len(alignment_core_hermit.unsat_classes),
        },
        "strict_bfo_mapping_path": None if strict_bfo_path is None else str(strict_bfo_path),
        "strict_bfo_mapping_sha256": strict_bfo_sha256,
        "strict_bfo_mapping": None
        if strict_bfo_result is None
        else {
            "product_key": strict_bfo_result.metadata.product_key,
            "stable_ontology_iri": strict_bfo_result.metadata.stable_ontology_iri,
            "governed_axiom_count": strict_bfo_result.governed_axiom_count,
            "logical_triple_count": strict_bfo_result.logical_triple_count,
            "ontology_declaration_triple_count": strict_bfo_result.ontology_declaration_triple_count,
            "import_triple_count": strict_bfo_result.import_triple_count,
            "metadata_annotation_count": strict_bfo_result.metadata_annotation_count,
            "total_triple_count": strict_bfo_result.total_triple_count,
            "subclass_axiom_count": strict_bfo_result.subclass_axiom_count,
            "equivalent_class_axiom_count": strict_bfo_result.equivalent_class_axiom_count,
            "direct_subproperty_axiom_count": strict_bfo_result.direct_subproperty_axiom_count,
            "property_chain_axiom_count": strict_bfo_result.property_chain_axiom_count,
            "domain_axiom_count": strict_bfo_result.domain_axiom_count,
            "range_axiom_count": strict_bfo_result.range_axiom_count,
            "union_expression_count": strict_bfo_result.union_target_count,
            "intersection_expression_count": strict_bfo_result.intersection_expression_count,
            "existential_restriction_count": strict_bfo_result.existential_restriction_count,
            "rdf_list_count": strict_bfo_result.rdf_list_count,
            "project_closure_governed_axiom_count": 48,
            "project_graph_triple_count": 195,
            "local_project_graph_triple_count": 194,
            "hermit_return_code": None if strict_bfo_hermit is None else strict_bfo_hermit.return_code,
            "hermit_result": "FAIL" if strict_bfo_hermit is None else "PASS" if strict_bfo_hermit.passed else "FAIL",
            "closure_triple_count": None if strict_bfo_hermit is None else strict_bfo_hermit.closure_triple_count,
            "named_unsat_count": None if strict_bfo_hermit is None else len(strict_bfo_hermit.unsat_classes),
        },
        "cco_extension_path": None if cco_extension_path is None else str(cco_extension_path),
        "cco_extension_sha256": cco_extension_sha256,
        "cco_extension": None
        if cco_extension_result is None
        else {
            "product_key": cco_extension_result.metadata.product_key,
            "stable_ontology_iri": cco_extension_result.metadata.stable_ontology_iri,
            "governed_axiom_count": cco_extension_result.governed_axiom_count,
            "cco_bearing_axiom_count": sum(
                axiom.target_category == "cco_bearing"
                for row in cco_extension_result.selected_rows
                for axiom in row.axioms
            ),
            "mixed_bfo_cco_axiom_count": sum(
                axiom.target_category == "mixed_bfo_cco"
                for row in cco_extension_result.selected_rows
                for axiom in row.axioms
            ),
            "logical_triple_count": cco_extension_result.logical_triple_count,
            "ontology_declaration_triple_count": cco_extension_result.ontology_declaration_triple_count,
            "import_triple_count": cco_extension_result.import_triple_count,
            "metadata_annotation_count": cco_extension_result.metadata_annotation_count,
            "total_triple_count": cco_extension_result.total_triple_count,
            "subclass_axiom_count": cco_extension_result.subclass_axiom_count,
            "equivalent_class_axiom_count": cco_extension_result.equivalent_class_axiom_count,
            "direct_subproperty_axiom_count": cco_extension_result.direct_subproperty_axiom_count,
            "property_chain_axiom_count": cco_extension_result.property_chain_axiom_count,
            "union_expression_count": cco_extension_result.union_target_count,
            "intersection_expression_count": cco_extension_result.intersection_expression_count,
            "existential_restriction_count": cco_extension_result.existential_restriction_count,
            "rdf_list_count": cco_extension_result.rdf_list_count,
            "project_closure_governed_axiom_count": 103,
            "project_graph_triple_count": 1128,
            "local_project_graph_triple_count": 1126,
            "hermit_return_code": None if cco_extension_hermit is None else cco_extension_hermit.return_code,
            "hermit_result": "FAIL" if cco_extension_hermit is None else "PASS" if cco_extension_hermit.passed else "FAIL",
            "closure_triple_count": None if cco_extension_hermit is None else cco_extension_hermit.closure_triple_count,
            "named_unsat_count": None if cco_extension_hermit is None else len(cco_extension_hermit.unsat_classes),
        },
        "worksheets_read": stats.worksheets_read,
        "active_axiom_rows": stats.active_axiom_rows,
        "mapped_rows": stats.mapped_rows,
        "blank_mapping_rows": stats.blank_mapping_rows,
        "class_mapping_rows": stats.class_mapping_rows,
        "object_property_mapping_rows": stats.object_property_mapping_rows,
        "domain_rows": stats.domain_rows,
        "range_rows": stats.range_rows,
        "property_chain_rows": stats.property_chain_rows,
        "governed_row_ids": stats.governed_row_id_count,
        "unique_row_ids": stats.unique_row_id_count,
        "processed_governed_rows": stats.processed_row_count,
        "identity_audit_rows": stats.identity_audit_row_count,
        "identity_count_reconciliation_passed": stats.identity_count_reconciliation_passed,
        "identity_row_id_set_reconciliation_passed": stats.identity_row_id_set_reconciliation_passed,
        "identity_location_reconciliation_passed": stats.identity_location_reconciliation_passed,
        "canonicalization_version": CANONICALIZATION_VERSION,
        "canonical_authoritative_axioms": sum(
            len(audit.authoritative_axioms) for audit in identity_audits
        ),
        "unique_canonical_authoritative_axioms": len(
            {
                axiom.canonical_axiom
                for audit in identity_audits
                for axiom in audit.authoritative_axioms
            }
        ),
        "product_dispositions": None
        if disposition_document is None
        else {
            "schema_version": disposition_document.schema_version,
            "product_order": list(disposition_document.product_order),
            **{
                field: getattr(disposition_document.summary, field)
                for field in disposition_document.summary.__dataclass_fields__
            },
        },
        "generated_ontology_triple_count": None if hermit is None else hermit.generated_triple_count,
        "candidate_closure_triple_count": None if hermit is None else hermit.closure_triple_count,
        "hermit_return_code": None if hermit is None else hermit.return_code,
        "hermit_result": "PASS" if hermit is not None and hermit.passed else "FAIL",
        "owl_nothing_count": None if hermit is None else hermit.owl_nothing_count,
        "named_unsat_count": None if hermit is None else len(hermit.unsat_classes),
        "named_unsat_set": [] if hermit is None else [str(value) for value in hermit.unsat_classes],
        "source_term_coverage": None
        if coverage is None
        else {
            "query_source_count": coverage.query_source_count,
            "query_unmapped_count": coverage.query_unmapped_count,
            "mapped_classes": len(coverage.mapped_classes),
            "unmapped_classes": len(coverage.unmapped_classes),
            "mapped_object_properties": len(coverage.mapped_object_properties),
            "unmapped_object_properties": len(coverage.unmapped_object_properties),
            "explicitly_listed_blank_mappings": len(coverage.explicit_blank_terms),
            "listed_only_in_domain_range_rows": len(coverage.property_typing_only_terms),
            "source_terms_absent_from_spreadsheet": len(coverage.absent_terms),
            "spreadsheet_subjects_not_found": len(coverage.spreadsheet_missing_subjects),
        },
        "coms_vs_pre_coms_legacy": None
        if comparison is None
        else {
            "mappings_present_in_both": len(comparison.both),
            "mappings_only_in_coms": len(comparison.coms_only),
            "mappings_only_in_legacy": len(comparison.legacy_only),
            "class_expression_differences": len(comparison.class_expression_differences),
            "object_property_differences": len(comparison.object_property_differences),
            "property_chain_differences": len(comparison.property_chain_differences),
            "domain_axioms_present_in_both": len(comparison.domain_both),
            "domain_axioms_only_in_coms": len(comparison.domain_coms_only),
            "domain_axioms_only_in_legacy": len(comparison.domain_legacy_only),
            "domain_target_differences": len(comparison.domain_differences),
            "range_axioms_present_in_both": len(comparison.range_both),
            "range_axioms_only_in_coms": len(comparison.range_coms_only),
            "range_axioms_only_in_legacy": len(comparison.range_legacy_only),
            "range_target_differences": len(comparison.range_differences),
        },
        "errors": errors,
        "runtime_seconds": round(elapsed_seconds, 3),
    }
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(payload, indent=2, sort_keys=True) + "\n", encoding="utf-8")


def main(argv: list[str] | None = None) -> int:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--input", required=True, help="COMS workbook input path.")
    parser.add_argument("--output", required=True, help="Generated TTL output path.")
    parser.add_argument("--report", required=True, help="Generation validation report path.")
    parser.add_argument(
        "--disposition-report",
        required=True,
        help="Generated COMS per-product disposition JSON path.",
    )
    parser.add_argument(
        "--alignment-core-output",
        required=True,
        help="Generated import-free SSN/SOSA alignment-core Turtle path.",
    )
    parser.add_argument(
        "--strict-bfo-output",
        required=True,
        help="Generated strict SSN/SOSA-to-BFO mapping Turtle path.",
    )
    parser.add_argument(
        "--cco-extension-output",
        required=True,
        help="Generated SSN/SOSA CCO-extension Turtle path.",
    )
    parser.add_argument(
        "--coverage-report",
        default="reports/coms-source-term-coverage.md",
        help="Source-term coverage report path.",
    )
    parser.add_argument(
        "--diff-report",
        default="reports/coms-vs-pre-coms-legacy-diff.md",
        help="COMS-vs-pre-COMS-legacy mapping diff report path.",
    )
    parser.add_argument(
        "--tmp-dir",
        default="/tmp/ssn-to-bfo-coms-generation",
        help="Temporary directory for candidate closure validation.",
    )
    parser.add_argument(
        "--report-workbook-path",
        help="Workbook path displayed in reports. Defaults to the actual --input path.",
    )
    parser.add_argument(
        "--report-output-path",
        help="Generated ontology path displayed in reports. Defaults to the actual --output path.",
    )
    parser.add_argument(
        "--report-disposition-path",
        help="Disposition path displayed in reports. Defaults to the actual --disposition-report path.",
    )
    parser.add_argument(
        "--report-alignment-core-path",
        help="Alignment-core path displayed in reports. Defaults to the actual --alignment-core-output path.",
    )
    parser.add_argument(
        "--report-strict-bfo-path",
        help="Strict-BFO path displayed in reports. Defaults to the actual --strict-bfo-output path.",
    )
    parser.add_argument(
        "--report-cco-extension-path",
        help="CCO-extension path displayed in reports. Defaults to the actual --cco-extension-output path.",
    )
    parser.add_argument(
        "--summary-json",
        help="Optional machine-readable generation summary path.",
    )
    args = parser.parse_args(argv)

    started = time.perf_counter()
    input_path = REPO_ROOT / args.input if not Path(args.input).is_absolute() else Path(args.input)
    output_path = REPO_ROOT / args.output if not Path(args.output).is_absolute() else Path(args.output)
    report_path = REPO_ROOT / args.report if not Path(args.report).is_absolute() else Path(args.report)
    disposition_report_path = REPO_ROOT / args.disposition_report if not Path(args.disposition_report).is_absolute() else Path(args.disposition_report)
    alignment_core_output_path = REPO_ROOT / args.alignment_core_output if not Path(args.alignment_core_output).is_absolute() else Path(args.alignment_core_output)
    strict_bfo_output_path = REPO_ROOT / args.strict_bfo_output if not Path(args.strict_bfo_output).is_absolute() else Path(args.strict_bfo_output)
    cco_extension_output_path = REPO_ROOT / args.cco_extension_output if not Path(args.cco_extension_output).is_absolute() else Path(args.cco_extension_output)
    coverage_report_path = REPO_ROOT / args.coverage_report if not Path(args.coverage_report).is_absolute() else Path(args.coverage_report)
    diff_report_path = REPO_ROOT / args.diff_report if not Path(args.diff_report).is_absolute() else Path(args.diff_report)
    report_workbook_path = Path(args.report_workbook_path) if args.report_workbook_path else input_path
    report_output_path = Path(args.report_output_path) if args.report_output_path else output_path
    report_disposition_path = Path(args.report_disposition_path) if args.report_disposition_path else disposition_report_path
    report_alignment_core_path = Path(args.report_alignment_core_path) if args.report_alignment_core_path else alignment_core_output_path
    report_strict_bfo_path = Path(args.report_strict_bfo_path) if args.report_strict_bfo_path else strict_bfo_output_path
    report_cco_extension_path = Path(args.report_cco_extension_path) if args.report_cco_extension_path else cco_extension_output_path
    summary_json_path = None
    if args.summary_json:
        summary_json_path = REPO_ROOT / args.summary_json if not Path(args.summary_json).is_absolute() else Path(args.summary_json)
    generation_timestamp = datetime.now(timezone.utc).isoformat(timespec="seconds")
    workbook_sha256 = sha256_file(input_path) if input_path.is_file() else "unavailable"
    generator_sha256 = sha256_file(Path(__file__).resolve())
    identity_module_sha256 = sha256_file(IDENTITY_MODULE)
    disposition_module_sha256 = sha256_file(DISPOSITION_MODULE)
    modular_products_module_sha256 = sha256_file(MODULAR_PRODUCTS_MODULE)
    publication_metadata_sha256 = sha256_file(PUBLICATION_METADATA)
    candidate_sha256 = "unavailable"
    disposition_sha256 = "unavailable"
    alignment_core_sha256 = "unavailable"
    strict_bfo_sha256 = "unavailable"
    cco_extension_sha256 = "unavailable"

    stats = WorkbookStats()
    resolver = Resolver()
    errors: list[str] = []
    hermit: HermitResult | None = None
    coverage: CoverageResult | None = None
    comparison: ComparisonResult | None = None
    normalized_rows: list[NormalizedRow] = []
    identity_audits: list[CanonicalRowAudit] = []
    disposition_document: DispositionDocument | None = None
    alignment_core_result: ModularProductResult | None = None
    alignment_core_hermit: HermitResult | None = None
    strict_bfo_result: ModularProductResult | None = None
    strict_bfo_hermit: HermitResult | None = None
    cco_extension_result: ModularProductResult | None = None
    cco_extension_hermit: HermitResult | None = None

    try:
        publication_metadata = load_metadata(PUBLICATION_METADATA)
        rows, stats = read_workbook(input_path)
        processed = validate_and_process_rows(rows, resolver, stats)
        identity_audits = [
            item.identity_audit
            for item in processed
            if item.identity_audit is not None
        ]
        disposition_document, _ = build_and_write_disposition_report(
            processed,
            disposition_report_path,
            RequiredInputHashes(
                workbook_sha256=workbook_sha256,
                generator_sha256=generator_sha256,
                row_identity_module_sha256=identity_module_sha256,
                disposition_module_sha256=disposition_module_sha256,
                publication_metadata_sha256=publication_metadata_sha256,
            ),
            publication_metadata,
        )
        disposition_sha256 = sha256_file(disposition_report_path)
        graph = generate_ontology(
            processed,
            output_path,
            publication_metadata,
            require_current_counts=True,
        )
        alignment_core_result, alignment_core_hermit = build_and_write_alignment_core(
            processed,
            identity_audits,
            disposition_document,
            publication_metadata,
            graph,
            alignment_core_output_path,
            Path(args.tmp_dir) / "alignment-core",
        )
        alignment_core_sha256 = sha256_file(alignment_core_output_path)
        strict_bfo_result, strict_bfo_hermit = build_and_write_strict_bfo_mapping(
            processed,
            identity_audits,
            disposition_document,
            publication_metadata,
            graph,
            alignment_core_result,
            alignment_core_output_path,
            strict_bfo_output_path,
            Path(args.tmp_dir) / "strict-bfo",
        )
        strict_bfo_sha256 = sha256_file(strict_bfo_output_path)
        cco_extension_result, cco_extension_hermit = build_and_write_cco_extension(
            processed,
            identity_audits,
            disposition_document,
            publication_metadata,
            graph,
            alignment_core_result,
            alignment_core_output_path,
            strict_bfo_result,
            strict_bfo_output_path,
            cco_extension_output_path,
            Path(args.tmp_dir) / "cco-extension",
        )
        cco_extension_sha256 = sha256_file(cco_extension_output_path)
        normalized_rows = normalized_axiom_rows(processed, graph)
        coverage = build_coverage(processed, [], coverage_report_path)
        graph.parse(output_path, format="turtle")
        hermit = run_candidate_hermit(output_path, Path(args.tmp_dir))
        comparison = compare_coms_to_legacy(output_path, diff_report_path, processed)
        if not hermit.passed:
            errors.append("candidate full local closure is not HermiT-clean")
    except GenerationError as exc:
        errors.append(str(exc))
    except Exception as exc:  # pragma: no cover - top-level reporting guard
        errors.append(f"unexpected generation failure: {type(exc).__name__}: {exc}")

    elapsed = time.perf_counter() - started
    if output_path.is_file():
        candidate_sha256 = sha256_file(output_path)
    write_generation_report(
        report_path,
        workbook_path=report_workbook_path,
        stats=stats,
        resolver=resolver,
        errors=errors,
        output_path=report_output_path,
        hermit=hermit,
        coverage=coverage,
        comparison=comparison,
        normalized_rows=normalized_rows,
        identity_audits=identity_audits,
        elapsed_seconds=elapsed,
        workbook_sha256=workbook_sha256,
        generator_sha256=generator_sha256,
        identity_module_sha256=identity_module_sha256,
        generation_timestamp=generation_timestamp,
        candidate_sha256=candidate_sha256,
        disposition_document=disposition_document,
        disposition_path=report_disposition_path,
        disposition_sha256=disposition_sha256,
        disposition_module_sha256=disposition_module_sha256,
        publication_metadata_sha256=publication_metadata_sha256,
        modular_products_module_sha256=modular_products_module_sha256,
        alignment_core_result=alignment_core_result,
        alignment_core_path=report_alignment_core_path,
        alignment_core_sha256=alignment_core_sha256,
        alignment_core_hermit=alignment_core_hermit,
        strict_bfo_result=strict_bfo_result,
        strict_bfo_path=report_strict_bfo_path,
        strict_bfo_sha256=strict_bfo_sha256,
        strict_bfo_hermit=strict_bfo_hermit,
        cco_extension_result=cco_extension_result,
        cco_extension_path=report_cco_extension_path,
        cco_extension_sha256=cco_extension_sha256,
        cco_extension_hermit=cco_extension_hermit,
    )
    if summary_json_path is not None:
        write_summary_json(
            summary_json_path,
            workbook_path=report_workbook_path,
            output_path=report_output_path,
            workbook_sha256=workbook_sha256,
            generator_sha256=generator_sha256,
            identity_module_sha256=identity_module_sha256,
            generation_timestamp=generation_timestamp,
            candidate_sha256=candidate_sha256,
            stats=stats,
            identity_audits=identity_audits,
            hermit=hermit,
            coverage=coverage,
            comparison=comparison,
            errors=errors,
            elapsed_seconds=elapsed,
            disposition_document=disposition_document,
            disposition_path=report_disposition_path,
            disposition_sha256=disposition_sha256,
            disposition_module_sha256=disposition_module_sha256,
            publication_metadata_sha256=publication_metadata_sha256,
            modular_products_module_sha256=modular_products_module_sha256,
            alignment_core_result=alignment_core_result,
            alignment_core_path=report_alignment_core_path,
            alignment_core_sha256=alignment_core_sha256,
            alignment_core_hermit=alignment_core_hermit,
            strict_bfo_result=strict_bfo_result,
            strict_bfo_path=report_strict_bfo_path,
            strict_bfo_sha256=strict_bfo_sha256,
            strict_bfo_hermit=strict_bfo_hermit,
            cco_extension_result=cco_extension_result,
            cco_extension_path=report_cco_extension_path,
            cco_extension_sha256=cco_extension_sha256,
            cco_extension_hermit=cco_extension_hermit,
        )

    print(f"Wrote {report_path}")
    if output_path.exists():
        print(f"Wrote {output_path}")
    if coverage_report_path.exists():
        print(f"Wrote {coverage_report_path}")
    if diff_report_path.exists():
        print(f"Wrote {diff_report_path}")
    if disposition_report_path.exists():
        print(f"Wrote {disposition_report_path}")
    if alignment_core_output_path.exists():
        print(f"Wrote {alignment_core_output_path}")
    if strict_bfo_output_path.exists():
        print(f"Wrote {strict_bfo_output_path}")
    if cco_extension_output_path.exists():
        print(f"Wrote {cco_extension_output_path}")
    print(f"Worksheets read: {', '.join(stats.worksheets_read) or 'none'}")
    print(f"Mapped rows: {stats.mapped_rows}")
    print(f"Blank mapping rows: {stats.blank_mapping_rows}")
    print(f"Class mapping rows: {stats.class_mapping_rows}")
    print(f"Object-property mapping rows: {stats.object_property_mapping_rows}")
    print(f"Domain rows: {stats.domain_rows}")
    print(f"Range rows: {stats.range_rows}")
    print(f"Property-chain rows: {stats.property_chain_rows}")
    print(f"Governed RowIDs: {stats.governed_row_id_count}")
    print(f"Unique RowIDs: {stats.unique_row_id_count}")
    print(f"Processed governed rows: {stats.processed_row_count}")
    print(f"Identity-audit rows: {stats.identity_audit_row_count}")
    print(
        "Canonical authoritative axioms: "
        f"{sum(len(audit.authoritative_axioms) for audit in identity_audits)}"
    )
    if disposition_document is not None:
        disposition_summary = disposition_document.summary
        print(f"Disposition target-neutral axioms: {disposition_summary.target_neutral_axiom_count}")
        print(f"Disposition BFO-bearing axioms: {disposition_summary.bfo_bearing_axiom_count}")
        print(f"Disposition CCO-bearing axioms: {disposition_summary.cco_bearing_axiom_count}")
        print(f"Disposition mixed BFO/CCO axioms: {disposition_summary.mixed_bfo_cco_axiom_count}")
    if alignment_core_result is not None:
        print(f"Alignment-core governed axioms: {alignment_core_result.governed_axiom_count}")
        print(f"Alignment-core logical triples: {alignment_core_result.logical_triple_count}")
        print(f"Alignment-core total triples: {alignment_core_result.total_triple_count}")
        print(f"Alignment-core SHA-256: {alignment_core_result.sha256}")
    if alignment_core_hermit is not None:
        print(f"Alignment-core source-closure HermiT: {'PASS' if alignment_core_hermit.passed else 'FAIL'}")
    if strict_bfo_result is not None:
        print(f"Strict-BFO governed axioms: {strict_bfo_result.governed_axiom_count}")
        print(f"Strict-BFO logical triples: {strict_bfo_result.logical_triple_count}")
        print(f"Strict-BFO total triples: {strict_bfo_result.total_triple_count}")
        print(f"Strict-BFO SHA-256: {strict_bfo_result.sha256}")
    if strict_bfo_hermit is not None:
        print(
            "Strict-BFO pinned merged CCO/BFO closure HermiT: "
            f"{'PASS' if strict_bfo_hermit.passed else 'FAIL'}"
        )
    if cco_extension_result is not None:
        print(f"CCO-extension governed axioms: {cco_extension_result.governed_axiom_count}")
        print(f"CCO-extension logical triples: {cco_extension_result.logical_triple_count}")
        print(f"CCO-extension total triples: {cco_extension_result.total_triple_count}")
        print(f"CCO-extension SHA-256: {cco_extension_result.sha256}")
    if cco_extension_hermit is not None:
        print(
            "CCO-extension pinned merged CCO/BFO closure HermiT: "
            f"{'PASS' if cco_extension_hermit.passed else 'FAIL'}"
        )
    if hermit is not None:
        print(f"Generated triple count: {hermit.generated_triple_count}")
        print(f"Candidate closure triple count: {hermit.closure_triple_count}")
        print(f"HermiT return code: {hermit.return_code}")
        print(f"owl:Nothing count: {hermit.owl_nothing_count}")
        print(f"Named unsat count: {len(hermit.unsat_classes)}")
        print(f"HermiT summary: {'PASS' if hermit.passed else 'FAIL'}")
    if coverage is not None:
        print(f"Mapped classes: {len(coverage.mapped_classes)}")
        print(f"Unmapped classes: {len(coverage.unmapped_classes)}")
        print(f"Mapped object properties: {len(coverage.mapped_object_properties)}")
        print(f"Unmapped object properties: {len(coverage.unmapped_object_properties)}")
    if comparison is not None:
        print(f"Mappings present in both: {len(comparison.both)}")
        print(f"COMS-only mappings: {len(comparison.coms_only)}")
        print(f"Pre-COMS legacy-only mappings: {len(comparison.legacy_only)}")
    if errors:
        print("Errors:")
        for error in errors:
            print(f"- {error}")
        return 1
    print("Summary: PASS")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
