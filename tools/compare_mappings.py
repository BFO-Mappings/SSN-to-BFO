#!/usr/bin/env python3
"""Compare SSN/SOSA mapping axioms in a spreadsheet against a Turtle file.

The comparison is intentionally conservative: spreadsheet rows and their
``OWL Axiom`` cells define the governed assertions, while the TTL extraction
ignores declarations and metadata triples unless they use mapping predicates.
"""

from __future__ import annotations

import argparse
import csv
import os
import re
import shlex
import subprocess
import sys
from collections import Counter, defaultdict
from dataclasses import dataclass, field
from pathlib import Path
from typing import Iterable

try:
    import openpyxl
except ModuleNotFoundError as exc:  # pragma: no cover - fail-fast runtime guard
    raise SystemExit("Missing dependency: openpyxl is required to inspect the spreadsheet.") from exc

try:
    from rdflib import BNode, Graph, Literal, URIRef
    from rdflib.namespace import OWL, RDF, RDFS, SKOS
except ModuleNotFoundError as exc:  # pragma: no cover - fail-fast runtime guard
    raise SystemExit("Missing dependency: rdflib is required to parse Turtle.") from exc


SCHEMA = "http://schema.org/"
DCTERMS = "http://purl.org/dc/terms/"

PREFIXES = {
    "bfo": "http://purl.obolibrary.org/obo/",
    "cco": "https://www.commoncoreontologies.org/",
    "obo": "http://purl.obolibrary.org/obo/",
    "owl": str(OWL),
    "rdf": str(RDF),
    "rdfs": str(RDFS),
    "schema": SCHEMA,
    "skos": str(SKOS),
    "sosa": "http://www.w3.org/ns/sosa/",
    "sampling": "http://www.w3.org/ns/sosa/sampling/",
    "sosa-rel": "http://www.w3.org/ns/sosa/sampling/",
    "ssn": "http://www.w3.org/ns/ssn/",
    "ssn-system": "http://www.w3.org/ns/ssn/systems/",
    "time": "http://www.w3.org/2006/time#",
    "xsd": "http://www.w3.org/2001/XMLSchema#",
}

SOURCE_NAMESPACES = (
    PREFIXES["sosa"],
    PREFIXES["ssn"],
    PREFIXES["ssn-system"],
)

TARGET_NAMESPACES = (
    PREFIXES["bfo"],
    PREFIXES["cco"],
)

MAPPING_PREDICATES = {
    str(RDFS.subClassOf): "rdfs:subClassOf",
    str(RDFS.subPropertyOf): "rdfs:subPropertyOf",
    str(OWL.equivalentClass): "owl:equivalentClass",
    str(OWL.equivalentProperty): "owl:equivalentProperty",
    str(OWL.propertyChainAxiom): "owl:propertyChainAxiom",
    str(SKOS.exactMatch): "skos:exactMatch",
    str(SKOS.closeMatch): "skos:closeMatch",
    str(SKOS.broadMatch): "skos:broadMatch",
    str(SKOS.narrowMatch): "skos:narrowMatch",
    str(SKOS.relatedMatch): "skos:relatedMatch",
    str(RDFS.seeAlso): "rdfs:seeAlso",
}

SUPPORTED_ISSUE_CATEGORIES = (
    "missing_in_ttl",
    "missing_in_spreadsheet",
    "target_mismatch",
    "relation_mismatch",
    "status_mismatch",
    "duplicate_mapping",
    "conflicting_mapping",
    "label_only_match",
    "prefix_or_iri_issue",
    "needs_human_review",
)

IGNORED_DECLARATION_TYPES = {
    str(OWL.Class),
    str(OWL.ObjectProperty),
    str(OWL.DatatypeProperty),
    str(OWL.AnnotationProperty),
    str(OWL.Ontology),
    str(RDF.Property),
}

STATUS_TERMS = {
    "tbd",
    "n/a",
    "na",
    "no mapping",
    "nomapping",
    "defer",
    "deferred",
    "reject",
    "rejected",
    "deprecated",
    "obsolete",
}

STRUCTURAL_NAMESPACES = (
    str(RDF),
    str(RDFS),
    str(OWL),
)

TOKEN_RE = re.compile(r"\b([A-Za-z][A-Za-z0-9_-]*:[A-Za-z_][A-Za-z0-9_.-]*)\b")
TRIPLE_RE = re.compile(
    r"^(?P<s>[A-Za-z][A-Za-z0-9_-]*:[^\s]+)\s+"
    r"(?P<p>(?:rdf|rdfs|owl|schema|skos):[^\s]+)\s+"
    r"(?P<o>.+?)\s*\.?$"
)

# Vocabulary used to construct OWL expressions, not domain mapping targets.
# This lets spreadsheet and TTL restriction expressions compare on their
# domain-relevant target IRIs, e.g. the property and filler in an OWL restriction.
STRUCTURAL_TARGET_IRIS = {
    str(RDF.type),
    str(OWL.Restriction),
    str(OWL.Class),
    str(OWL.ObjectProperty),
    str(OWL.onProperty),
    str(OWL.someValuesFrom),
    str(OWL.allValuesFrom),
    str(OWL.hasValue),
    str(OWL.propertyChainAxiom),
}


def comparable_target_iris(iris: Iterable[str]) -> tuple[str, ...]:
    return tuple(sorted(set(iri for iri in iris if iri not in STRUCTURAL_TARGET_IRIS)))


@dataclass(frozen=True)
class ResolvedToken:
    token: str
    iri: str | None
    issue: str | None = None


@dataclass
class ExpectedAssertion:
    sheet: str
    row: int
    source_token: str
    source_iri: str | None
    source_issue: str | None
    predicate_iri: str | None
    predicate_label: str
    target_text: str
    target_iris: tuple[str, ...]
    unresolved_tokens: tuple[str, ...]
    axiom: str
    notes: str
    status_text: str | None = None

    @property
    def key(self) -> tuple[str | None, str | None, tuple[str, ...]]:
        return (self.source_iri, self.predicate_iri, tuple(sorted(self.target_iris)))


@dataclass
class SpreadsheetRow:
    sheet: str
    row: int
    source_token: str
    source_iri: str | None
    source_issue: str | None
    axiom: str
    notes: str
    expected: list[ExpectedAssertion] = field(default_factory=list)
    status_text: str | None = None


@dataclass
class SheetSchema:
    sheet: str
    header_row: int
    source_column: str
    owl_axiom_column: str
    note_columns: tuple[str, ...]
    status_columns: tuple[str, ...]
    candidate_mapping_rows: int


@dataclass
class TtlAssertion:
    source_iri: str
    predicate_iri: str
    predicate_label: str
    target_iris: tuple[str, ...]
    target_summary: str
    line: int | None

    @property
    def key(self) -> tuple[str, str, tuple[str, ...]]:
        return (self.source_iri, self.predicate_iri, tuple(sorted(self.target_iris)))


@dataclass
class Issue:
    category: str
    sheet: str
    row: int | str
    source: str
    source_iri: str
    ttl_predicate: str
    spreadsheet_relation: str
    ttl_target: str
    spreadsheet_target: str
    ttl_line: int | str
    recommended_action: str


@dataclass
class GitMetadata:
    branch: str
    commit: str
    untracked_audit_files: bool
    audit_status_entries: tuple[str, ...]


def normalize_text(value: object) -> str:
    if value is None:
        return ""
    return re.sub(r"\s+", " ", str(value).strip())


def normalize_key(value: str) -> str:
    return re.sub(r"[^a-z0-9]+", "", value.lower())


def local_name(iri: str) -> str:
    for sep in ("#", "/"):
        if sep in iri:
            iri = iri.rsplit(sep, 1)[-1]
    return iri


def display_iri(iri: str | None) -> str:
    if not iri:
        return ""
    for prefix, ns in sorted(PREFIXES.items(), key=lambda item: len(item[1]), reverse=True):
        if iri.startswith(ns):
            return f"{prefix}:{iri[len(ns):]}"
    return iri


def display_iris(iris: Iterable[str]) -> str:
    return "; ".join(display_iri(iri) for iri in sorted(set(iris)))


def safe_cell(value: object) -> str:
    return normalize_text(value).replace("|", "\\|")


def column_label(column_index: int, header: str) -> str:
    letter = openpyxl.utils.get_column_letter(column_index)
    return f"{letter} (`{header}`)"


def current_command() -> str:
    parts = []
    pythonpath = os.environ.get("PYTHONPATH")
    if pythonpath:
        parts.append(f"PYTHONPATH={shlex.quote(pythonpath)}")
    parts.append(shlex.quote(sys.executable))
    parts.extend(shlex.quote(arg) for arg in sys.argv)
    return " ".join(parts)


def portable_command() -> str:
    return (
        'python tools/compare_mappings.py --ttl legacy/SSN2BFO-pre-COMS.ttl '
        '--spreadsheet "Current_SOSA-SSN to BFO-CCO.xlsx" '
        "--output-md reports/mapping-consistency-audit.md "
        "--output-csv reports/mapping-consistency-audit.csv"
    )


def git_metadata(repo_root: Path) -> GitMetadata:
    def run_git(*args: str) -> str:
        try:
            return subprocess.check_output(["git", *args], cwd=repo_root, text=True, stderr=subprocess.DEVNULL).strip()
        except (subprocess.CalledProcessError, FileNotFoundError):
            return "unavailable"

    branch = run_git("branch", "--show-current")
    commit = run_git("rev-parse", "HEAD")
    status = run_git(
        "status",
        "--porcelain",
        "--untracked-files=all",
        "--",
        "tools/compare_mappings.py",
        "reports/mapping-consistency-audit.md",
        "reports/mapping-consistency-audit.csv",
    )
    entries = tuple(line for line in status.splitlines() if line)
    return GitMetadata(
        branch=branch or "unavailable",
        commit=commit or "unavailable",
        untracked_audit_files=any(entry.startswith("??") for entry in entries),
        audit_status_entries=entries,
    )


def parse_graph(path: Path) -> Graph:
    graph = Graph()
    graph.parse(path, format="turtle")
    return graph


def parse_optional_graph(path: Path) -> Graph:
    if not path.exists():
        return Graph()
    return parse_graph(path)


def build_label_index(repo_root: Path) -> tuple[dict[tuple[str, str], set[str]], dict[str, str]]:
    label_index: dict[tuple[str, str], set[str]] = defaultdict(set)
    iri_labels: dict[str, str] = {}
    for rel in ("imports/cco.ttl", "imports/ssn.ttl", "imports/ssn-systems.ttl"):
        graph = parse_optional_graph(repo_root / rel)
        for subject, _, label in graph.triples((None, RDFS.label, None)):
            if not isinstance(subject, URIRef) or not isinstance(label, Literal):
                continue
            iri = str(subject)
            label_text = str(label)
            iri_labels.setdefault(iri, label_text)
            for prefix, ns in PREFIXES.items():
                if iri.startswith(ns):
                    label_index[(prefix, normalize_key(label_text))].add(iri)
                    label_index[(prefix, normalize_key(label_text.replace(" ", "_")))].add(iri)
                    label_index[(prefix, normalize_key(local_name(iri)))].add(iri)
    return label_index, iri_labels


class Resolver:
    def __init__(self, label_index: dict[tuple[str, str], set[str]]):
        self.label_index = label_index

    def resolve(self, token: str) -> ResolvedToken:
        token = token.strip().strip(".,;()[]")
        if ":" not in token:
            return ResolvedToken(token, None, "not_a_curie")
        prefix, local = token.split(":", 1)
        prefix_l = prefix.lower()
        if prefix_l not in PREFIXES:
            return ResolvedToken(token, None, f"unresolved_prefix:{prefix}")
        ns = PREFIXES[prefix_l]
        if prefix_l in {"bfo", "cco"} and not re.match(r"^(BFO_|IAO_|RO_|ont)\d+", local):
            candidates = sorted(self.label_index.get((prefix_l, normalize_key(local)), set()))
            if len(candidates) == 1:
                return ResolvedToken(token, candidates[0])
            if len(candidates) > 1:
                return ResolvedToken(token, None, f"ambiguous_label:{token}")
            return ResolvedToken(token, None, f"unresolved_label:{token}")
        return ResolvedToken(token, ns + local)


def relation_from_token(token: str) -> tuple[str | None, str]:
    normalized = token.strip()
    if normalized == "equivalentTo":
        return str(OWL.equivalentClass), "owl:equivalentClass"
    if normalized == "subClassOf":
        return str(RDFS.subClassOf), "rdfs:subClassOf"
    if normalized == "subPropertyOf":
        return str(RDFS.subPropertyOf), "rdfs:subPropertyOf"
    resolved = Resolver({}).resolve(normalized)
    if resolved.iri in MAPPING_PREDICATES:
        return resolved.iri, MAPPING_PREDICATES[resolved.iri]
    return resolved.iri, normalized


def detect_status(*values: str) -> str | None:
    for value in values:
        text = normalize_text(value).lower().strip(" .;:-")
        if not text:
            continue
        if text in STATUS_TERMS or normalize_key(text) in {normalize_key(term) for term in STATUS_TERMS}:
            return text
        match = re.search(
            r"\b(?:status|review|decision)\s*[:=-]\s*(tbd|n/a|no mapping|defer(?:red)?|reject(?:ed)?|deprecated|obsolete)\b",
            text,
        )
        if match:
            return match.group(1)
    return None


def split_top_level(text: str, delimiter: str) -> list[str]:
    parts: list[str] = []
    depth_square = 0
    depth_paren = 0
    start = 0
    for idx, char in enumerate(text):
        if char == "[":
            depth_square += 1
        elif char == "]" and depth_square:
            depth_square -= 1
        elif char == "(":
            depth_paren += 1
        elif char == ")" and depth_paren:
            depth_paren -= 1
        elif char == delimiter and depth_square == 0 and depth_paren == 0:
            parts.append(text[start:idx].strip())
            start = idx + 1
    parts.append(text[start:].strip())
    return [part for part in parts if part]


def split_axiom_statements(axiom: str) -> list[str]:
    statements: list[str] = []
    for part in re.split(r"\n+", axiom):
        part = part.strip()
        if not part:
            continue
        if " . " in part:
            statements.extend(p.strip() for p in part.split(" . ") if p.strip())
        else:
            statements.append(part)
    return statements


def extract_target_tokens(target_text: str, source_token: str) -> list[str]:
    tokens = []
    source_norm = source_token.lower()
    for token in TOKEN_RE.findall(target_text):
        if token.lower() == source_norm:
            continue
        tokens.append(token)
    return tokens


def parse_spreadsheet(path: Path, resolver: Resolver) -> tuple[list[SpreadsheetRow], list[SheetSchema]]:
    workbook = openpyxl.load_workbook(path, read_only=True, data_only=False)
    schemas: list[SheetSchema] = []
    rows: list[SpreadsheetRow] = []
    for ws in workbook.worksheets:
        headers = [normalize_text(ws.cell(row=1, column=col).value) for col in range(1, ws.max_column + 1)]
        header_map = {header.lower(): idx + 1 for idx, header in enumerate(headers)}
        if not any(headers):
            continue
        if "iri" not in header_map or "owl axiom" not in header_map:
            continue
        status_cols = [
            idx + 1
            for idx, header in enumerate(headers)
            if any(marker in header.lower() for marker in ("status", "review", "decision"))
        ]
        note_cols = [
            header_map[name]
            for name in ("definition", "bfo definition", "natural language owl", "reasoning", "shacl")
            if name in header_map
        ]
        candidate_rows = 0
        for row_idx in range(2, ws.max_row + 1):
            source_token = normalize_text(ws.cell(row=row_idx, column=header_map["iri"]).value)
            axiom = normalize_text(ws.cell(row=row_idx, column=header_map["owl axiom"]).value)
            if not source_token and not axiom:
                continue
            candidate_rows += 1
            source = resolver.resolve(source_token) if source_token else ResolvedToken("", None, "missing_source")
            notes = " | ".join(
                normalize_text(ws.cell(row=row_idx, column=col).value)
                for col in note_cols
                if normalize_text(ws.cell(row=row_idx, column=col).value)
            )
            row = SpreadsheetRow(
                sheet=ws.title,
                row=row_idx,
                source_token=source_token,
                source_iri=source.iri,
                source_issue=source.issue,
                axiom=axiom,
                notes=notes,
                status_text=detect_status(
                    axiom if normalize_key(axiom) in {normalize_key(term) for term in STATUS_TERMS} else "",
                    *[normalize_text(ws.cell(row=row_idx, column=col).value) for col in status_cols],
                ),
            )
            if axiom:
                row.expected.extend(parse_expected_assertions(row, resolver))
            rows.append(row)
        schemas.append(
            SheetSchema(
                sheet=ws.title,
                header_row=1,
                source_column=column_label(header_map["iri"], "IRI"),
                owl_axiom_column=column_label(header_map["owl axiom"], "OWL Axiom"),
                note_columns=tuple(column_label(col, headers[col - 1]) for col in note_cols),
                status_columns=tuple(column_label(col, headers[col - 1]) for col in status_cols),
                candidate_mapping_rows=candidate_rows,
            )
        )
    if not schemas:
        raise SystemExit("Spreadsheet schema blocker: no sheet with both 'IRI' and 'OWL Axiom' columns was detected.")
    return rows, schemas


def parse_expected_assertions(row: SpreadsheetRow, resolver: Resolver) -> list[ExpectedAssertion]:
    expected: list[ExpectedAssertion] = []

    def add_assertion(predicate_token: str, target_text: str) -> None:
        predicate_iri, predicate_label = relation_from_token(predicate_token)
        if predicate_iri not in MAPPING_PREDICATES:
            return
        target_iris: list[str] = []
        unresolved: list[str] = []
        for token in extract_target_tokens(target_text, row.source_token):
            resolved = resolver.resolve(token)
            if resolved.iri:
                if resolved.iri not in STRUCTURAL_TARGET_IRIS:
                    target_iris.append(resolved.iri)
            else:
                unresolved.append(f"{token} ({resolved.issue})")
        expected.append(
            ExpectedAssertion(
                sheet=row.sheet,
                row=row.row,
                source_token=row.source_token,
                source_iri=row.source_iri,
                source_issue=row.source_issue,
                predicate_iri=predicate_iri,
                predicate_label=predicate_label,
                target_text=target_text,
                target_iris=tuple(sorted(set(target_iris))),
                unresolved_tokens=tuple(unresolved),
                axiom=row.axiom,
                notes=row.notes,
                status_text=row.status_text,
            )
        )

    for statement in split_axiom_statements(row.axiom):
        statement = statement.strip().rstrip(".").strip()
        triple_match = TRIPLE_RE.match(statement)
        if triple_match:
            predicate_token = triple_match.group("p")
            target_text = triple_match.group("o")
            parts = split_top_level(target_text, ";")
            if parts:
                add_assertion(predicate_token, parts[0])
                for continuation in parts[1:]:
                    cont_match = re.match(r"^(?P<p>(?:rdf|rdfs|owl|schema|skos):[^\s]+)\s+(?P<o>.+)$", continuation)
                    if cont_match:
                        add_assertion(cont_match.group("p"), cont_match.group("o"))
        else:
            keyword_match = re.match(r"^(subClassOf|subPropertyOf|equivalentTo)\s+(.+)$", statement)
            if not keyword_match:
                expected.append(
                    ExpectedAssertion(
                        sheet=row.sheet,
                        row=row.row,
                        source_token=row.source_token,
                        source_iri=row.source_iri,
                        source_issue=row.source_issue,
                        predicate_iri=None,
                        predicate_label="unparsed",
                        target_text=statement,
                        target_iris=(),
                        unresolved_tokens=(),
                        axiom=row.axiom,
                        notes=row.notes,
                        status_text=row.status_text,
                    )
                )
                continue
            add_assertion(keyword_match.group(1), keyword_match.group(2))
    return expected


def ttl_subject_line_index(path: Path) -> dict[str, int]:
    line_index: dict[str, int] = {}
    header_re = re.compile(r"^###\s+(\S+)")
    subject_re = re.compile(r"^<([^>]+)>\s+")
    pending: tuple[str, int] | None = None
    for idx, line in enumerate(path.read_text(encoding="utf-8").splitlines(), start=1):
        header = header_re.match(line)
        if header:
            pending = (header.group(1), idx)
            continue
        subject = subject_re.match(line)
        if subject:
            iri = subject.group(1)
            if pending and pending[0] == iri:
                line_index.setdefault(iri, pending[1])
            else:
                line_index.setdefault(iri, idx)
            pending = None
    return line_index


def collect_uri_refs(graph: Graph, node: object, seen: set[object] | None = None) -> set[str]:
    if seen is None:
        seen = set()
    if node in seen:
        return set()
    seen.add(node)
    if isinstance(node, URIRef):
        return {str(node)} if not str(node).startswith(STRUCTURAL_NAMESPACES) else set()
    if isinstance(node, Literal):
        return set()
    collected: set[str] = set()
    if isinstance(node, BNode):
        for _, predicate, obj in graph.triples((node, None, None)):
            if isinstance(obj, URIRef) and not str(obj).startswith(STRUCTURAL_NAMESPACES):
                collected.add(str(obj))
            elif isinstance(obj, BNode):
                collected.update(collect_uri_refs(graph, obj, seen))
            if predicate == RDF.first:
                collected.update(collect_uri_refs(graph, obj, seen))
    return collected


def extract_ttl_assertions(path: Path, spreadsheet_sources: set[str]) -> tuple[list[TtlAssertion], Counter]:
    graph = parse_graph(path)
    line_index = ttl_subject_line_index(path)
    assertions: list[TtlAssertion] = []
    ignored = Counter()
    for subject, predicate, obj in graph:
        if not isinstance(subject, URIRef):
            ignored["blank_subject"] += 1
            continue
        subject_iri = str(subject)
        predicate_iri = str(predicate)
        object_iri = str(obj) if isinstance(obj, URIRef) else None
        if predicate == RDF.type and object_iri in IGNORED_DECLARATION_TYPES:
            ignored["declaration"] += 1
            continue
        if predicate_iri not in MAPPING_PREDICATES:
            ignored["non_mapping_predicate"] += 1
            continue
        is_source = subject_iri in spreadsheet_sources or subject_iri.startswith(SOURCE_NAMESPACES)
        if not is_source:
            ignored["non_source_subject"] += 1
            continue
        target_iris = comparable_target_iris(collect_uri_refs(graph, obj))
        assertions.append(
            TtlAssertion(
                source_iri=subject_iri,
                predicate_iri=predicate_iri,
                predicate_label=MAPPING_PREDICATES[predicate_iri],
                target_iris=tuple(sorted(target_iris)),
                target_summary=display_iris(target_iris) if target_iris else normalize_text(obj),
                line=line_index.get(subject_iri),
            )
        )
    return sorted(assertions, key=lambda a: (a.source_iri, a.predicate_iri, a.target_iris)), ignored


def labelish_overlap(left: Iterable[str], right: Iterable[str]) -> bool:
    left_set = set(left)
    right_set = set(right)
    if left_set & right_set:
        return False
    left_keys = {normalize_key(local_name(iri)) for iri in left}
    right_keys = {normalize_key(local_name(iri)) for iri in right}
    return bool(left_keys & right_keys)


def make_issue(
    category: str,
    expected: ExpectedAssertion | None,
    ttl: TtlAssertion | None,
    action: str,
) -> Issue:
    source_iri = expected.source_iri if expected else (ttl.source_iri if ttl else "")
    source = expected.source_token if expected else display_iri(source_iri)
    return Issue(
        category=category,
        sheet=expected.sheet if expected else "",
        row=expected.row if expected else "",
        source=source,
        source_iri=source_iri or "",
        ttl_predicate=ttl.predicate_label if ttl else "",
        spreadsheet_relation=expected.predicate_label if expected else "",
        ttl_target=ttl.target_summary if ttl else "",
        spreadsheet_target=(
            display_iris(expected.target_iris)
            + (f"; unresolved: {'; '.join(expected.unresolved_tokens)}" if expected and expected.unresolved_tokens else "")
            if expected
            else ""
        ),
        ttl_line=ttl.line if ttl and ttl.line else "",
        recommended_action=action,
    )


def compare(rows: list[SpreadsheetRow], ttl_assertions: list[TtlAssertion]) -> tuple[list[Issue], set[tuple]]:
    issues: list[Issue] = []
    matched_ttl_keys: set[tuple] = set()
    ttl_by_source: dict[str, list[TtlAssertion]] = defaultdict(list)
    ttl_by_key: dict[tuple, list[TtlAssertion]] = defaultdict(list)
    for assertion in ttl_assertions:
        ttl_by_source[assertion.source_iri].append(assertion)
        ttl_by_key[assertion.key].append(assertion)

    for row in rows:
        if row.source_issue:
            issues.append(
                make_issue(
                    "prefix_or_iri_issue",
                    row.expected[0]
                    if row.expected
                    else ExpectedAssertion(
                        row.sheet,
                        row.row,
                        row.source_token,
                        row.source_iri,
                        row.source_issue,
                        None,
                        "",
                        row.axiom,
                        (),
                        (),
                        row.axiom,
                        row.notes,
                        row.status_text,
                    ),
                    None,
                    f"Resolve spreadsheet source IRI issue before comparing: {row.source_issue}.",
                )
            )
            continue
        if row.status_text and row.source_iri in ttl_by_source:
            for ttl in ttl_by_source[row.source_iri]:
                issues.append(
                    make_issue(
                        "status_mismatch",
                        row.expected[0] if row.expected else None,
                        ttl,
                        "Review TTL mapping because spreadsheet text indicates a non-final or deferred status.",
                    )
                )
        for expected in row.expected:
            if expected.predicate_iri is None:
                issues.append(
                    make_issue(
                        "needs_human_review",
                        expected,
                        None,
                        "Spreadsheet OWL Axiom could not be parsed into a supported comparison assertion.",
                    )
                )
                continue
            if expected.unresolved_tokens:
                issues.append(
                    make_issue(
                        "prefix_or_iri_issue",
                        expected,
                        None,
                        "Resolve unresolved spreadsheet target prefixes or label aliases before applying a mechanical correction.",
                    )
                )
                continue
            candidates = ttl_by_source.get(expected.source_iri or "", [])
            exact = [ttl for ttl in candidates if ttl.key == expected.key]
            if exact:
                matched_ttl_keys.add(exact[0].key)
                continue
            same_relation = [ttl for ttl in candidates if ttl.predicate_iri == expected.predicate_iri]
            if same_relation:
                best = same_relation[0]
                category = "label_only_match" if labelish_overlap(expected.target_iris, best.target_iris) else "target_mismatch"
                issues.append(
                    make_issue(
                        category,
                        expected,
                        best,
                        "Compare the spreadsheet OWL Axiom and TTL expression; align target IRIs only after human review.",
                    )
                )
                continue
            same_target = [
                ttl
                for ttl in candidates
                if set(ttl.target_iris) == set(expected.target_iris)
                or (expected.target_iris and set(expected.target_iris).issubset(set(ttl.target_iris)))
            ]
            if same_target:
                issues.append(
                    make_issue(
                        "relation_mismatch",
                        expected,
                        same_target[0],
                        "Review whether the TTL relation is intentionally stronger/weaker than the spreadsheet relation.",
                    )
                )
                continue
            issues.append(
                make_issue(
                    "missing_in_ttl",
                    expected,
                    None,
                    "Add or revise the TTL mapping only after confirming the spreadsheet row is authoritative.",
                )
            )

    expected_sources = {row.source_iri for row in rows if row.source_iri}
    expected_keys = {expected.key for row in rows for expected in row.expected if expected.predicate_iri and not expected.unresolved_tokens}
    for ttl in ttl_assertions:
        if ttl.key in expected_keys:
            continue
        if ttl.source_iri not in expected_sources:
            issues.append(
                make_issue(
                    "missing_in_spreadsheet",
                    None,
                    ttl,
                    "Add a spreadsheet row or document why this TTL mapping is outside spreadsheet governance.",
                )
            )
        elif not any(
            issue.source_iri == ttl.source_iri
            and issue.ttl_predicate == ttl.predicate_label
            and issue.ttl_target == ttl.target_summary
            for issue in issues
        ):
            issues.append(
                make_issue(
                    "missing_in_spreadsheet",
                    None,
                    ttl,
                    "Reconcile this extra TTL mapping with the source spreadsheet row for the same term.",
                )
            )

    grouped = defaultdict(list)
    for ttl in ttl_assertions:
        grouped[ttl.key].append(ttl)
    for duplicates in grouped.values():
        if len(duplicates) > 1:
            issues.append(
                make_issue(
                    "duplicate_mapping",
                    None,
                    duplicates[0],
                    "Remove redundant TTL assertion only if it is not required by the serialization structure.",
                )
            )
    return sorted(issues, key=lambda i: (str(i.row).zfill(6), i.category, i.source, i.ttl_predicate, i.ttl_target)), matched_ttl_keys


def write_csv(path: Path, issues: list[Issue]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    fields = [
        "issue_id",
        "category",
        "sheet",
        "spreadsheet_row",
        "source_term",
        "source_iri",
        "ttl_predicate",
        "spreadsheet_relation",
        "ttl_target",
        "spreadsheet_target",
        "ttl_line",
        "recommended_action",
    ]
    with path.open("w", newline="", encoding="utf-8") as handle:
        writer = csv.DictWriter(handle, fieldnames=fields, lineterminator="\n")
        writer.writeheader()
        for idx, issue in enumerate(issues, start=1):
            writer.writerow(
                {
                    "issue_id": f"ISSUE-{idx:04d}",
                    "category": issue.category,
                    "sheet": issue.sheet,
                    "spreadsheet_row": issue.row,
                    "source_term": issue.source,
                    "source_iri": issue.source_iri,
                    "ttl_predicate": issue.ttl_predicate,
                    "spreadsheet_relation": issue.spreadsheet_relation,
                    "ttl_target": issue.ttl_target,
                    "spreadsheet_target": issue.spreadsheet_target,
                    "ttl_line": issue.ttl_line,
                    "recommended_action": issue.recommended_action,
                }
            )


def write_markdown(
    path: Path,
    *,
    ttl_path: Path,
    spreadsheet_path: Path,
    exact_command: str,
    portable_command_text: str,
    schemas: list[SheetSchema],
    git_meta: GitMetadata,
    rows: list[SpreadsheetRow],
    ttl_assertions: list[TtlAssertion],
    issues: list[Issue],
    matched_ttl_keys: set[tuple],
    ignored: Counter,
) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    counts = Counter(issue.category for issue in issues)
    expected_assertions = [expected for row in rows for expected in row.expected]
    exact_row_matches = 0
    for row in rows:
        comparable = [e for e in row.expected if e.predicate_iri and not e.unresolved_tokens]
        if comparable and all(e.key in matched_ttl_keys for e in comparable):
            exact_row_matches += 1

    skipped_rows = [
        row
        for row in rows
        if row.source_issue or not row.expected or any(e.predicate_iri is None or e.unresolved_tokens for e in row.expected)
    ]

    lines: list[str] = []
    lines.append("# Mapping Consistency Audit")
    lines.append("")
    lines.append("## Files Inspected")
    lines.append(f"- TTL mapping file: `{ttl_path}`")
    lines.append(f"- Spreadsheet mapping source: `{spreadsheet_path}`")
    lines.append(f"- Sheets inspected: {', '.join(f'`{schema.sheet}`' for schema in schemas)}")
    lines.append("")
    lines.append("## Git Context")
    lines.append(f"- Current branch: `{git_meta.branch}`")
    lines.append(f"- Current commit: `{git_meta.commit}`")
    lines.append(
        f"- Working tree had untracked audit files at report generation time: {'yes' if git_meta.untracked_audit_files else 'no'}"
    )
    if git_meta.audit_status_entries:
        lines.append("- Audit file status entries at report generation time:")
        for entry in git_meta.audit_status_entries:
            lines.append(f"  - `{entry}`")
    else:
        lines.append("- Audit file status entries at report generation time: none")
    lines.append("")
    lines.append("## Detected Spreadsheet Schema")
    lines.append(
        "| Sheet | Header Row | Source/IRI Column | OWL Axiom Column | Comment/Notes Columns | Status/Review Columns | Candidate Mapping Rows |"
    )
    lines.append("| --- | ---: | --- | --- | --- | --- | ---: |")
    for schema in schemas:
        lines.append(
            "| "
            + " | ".join(
                [
                    safe_cell(schema.sheet),
                    str(schema.header_row),
                    safe_cell(schema.source_column),
                    safe_cell(schema.owl_axiom_column),
                    safe_cell("; ".join(schema.note_columns) if schema.note_columns else "none detected"),
                    safe_cell("; ".join(schema.status_columns) if schema.status_columns else "none detected"),
                    str(schema.candidate_mapping_rows),
                ]
            )
            + " |"
        )
    lines.append("")
    lines.append("## Comparison Method")
    lines.append(
        "- Spreadsheet rows were treated as authoritative only when a sheet had both `IRI` and `OWL Axiom` columns."
    )
    lines.append(
        "- TTL assertions were considered candidate mapping assertions only when they used recognized mapping predicates and had SSN/SOSA source subjects or spreadsheet-governed source subjects."
    )
    lines.append(
        "- Declarations, imports, labels, comments, and non-mapping metadata triples were ignored for mismatch classification."
    )
    lines.append(
        "- Prefixes were expanded to full IRIs. CCO and BFO label-style CURIEs in the spreadsheet were resolved through labels in `imports/cco.ttl` where unique."
    )
    lines.append(
        "- Blank-node OWL class expressions were summarized by the named IRIs they contain; these require human review before any ontology edit."
    )
    lines.append("")
    lines.append("## TTL Extraction Criteria")
    lines.append(
        "- Candidate mapping assertions: triples whose predicate is one of the recognized mapping predicates, whose subject is an SSN/SOSA source IRI or a spreadsheet-governed source IRI, and whose object or blank-node expression can be summarized for comparison."
    )
    lines.append(
        "- Supporting ontology/context triples: blank-node triples inside OWL restrictions, intersections, unions, and property chains. These are traversed only to collect named IRIs for a candidate mapping expression."
    )
    lines.append(
        "- Ignored metadata/declaration triples: ontology imports, `rdf:type` declarations for classes/properties/ontologies, labels, comments, definitions, and other non-mapping predicates."
    )
    lines.append(
        "- Declarations are excluded because they only state entity kind, not mapping intent. Imports are excluded because they establish context, not source-to-target mapping rows. Labels/comments/definitions are excluded because the audit compares IRIs and asserted relations rather than relying on prose. Blank-subject triples are excluded as standalone mappings because they are expression structure without their owning source subject. Non-mapping predicates are excluded to keep supporting ontology context from being reported as spreadsheet-governed mappings."
    )
    lines.append("")
    lines.append("## Exact Command Used")
    lines.append("")
    lines.append("```bash")
    lines.append(exact_command)
    lines.append("```")
    lines.append("")
    lines.append("## Portable Command Example")
    lines.append("")
    lines.append("```bash")
    lines.append(portable_command_text)
    lines.append("```")
    lines.append("")
    lines.append("## Summary")
    lines.append(f"- Total spreadsheet mapping rows: {len(rows)}")
    lines.append(f"- Total spreadsheet expected assertions parsed: {len(expected_assertions)}")
    lines.append(f"- Total TTL candidate mapping assertions: {len(ttl_assertions)}")
    lines.append(f"- Exact spreadsheet row matches: {exact_row_matches}")
    lines.append(f"- Exact assertion matches: {len(matched_ttl_keys)}")
    lines.append(f"- Total issues: {len(issues)}")
    lines.append("")
    lines.append("## Issues by Category")
    for category in SUPPORTED_ISSUE_CATEGORIES:
        lines.append(f"- `{category}`: {counts.get(category, 0)}")
    lines.append("")
    lines.append(
        "Note: not all supported issue categories necessarily appear in this run; zero-count categories are still supported by the audit taxonomy."
    )
    lines.append("")
    lines.append("## Ignored TTL Triples")
    for key, value in sorted(ignored.items()):
        lines.append(f"- `{key}`: {value}")
    lines.append("")
    lines.append("## Skipped or Partially Parsed Rows")
    if skipped_rows:
        lines.append("| Sheet | Row | Source | Reason |")
        lines.append("| --- | ---: | --- | --- |")
        for row in skipped_rows:
            reasons = []
            if row.source_issue:
                reasons.append(row.source_issue)
            if not row.expected:
                reasons.append("no parsed expected assertions")
            for expected in row.expected:
                if expected.predicate_iri is None:
                    reasons.append("unparsed OWL Axiom")
                if expected.unresolved_tokens:
                    reasons.append("unresolved tokens: " + "; ".join(expected.unresolved_tokens))
            lines.append(f"| {safe_cell(row.sheet)} | {row.row} | {safe_cell(row.source_token)} | {safe_cell('; '.join(sorted(set(reasons))))} |")
    else:
        lines.append("- None.")
    lines.append("")
    lines.append("## Detailed Issues")
    lines.append(
        "| Issue ID | Category | Sheet | Row | Source | Source IRI | TTL Predicate | Spreadsheet Relation | TTL Target | Spreadsheet Target | TTL Line | Recommended Action |"
    )
    lines.append("| --- | --- | --- | ---: | --- | --- | --- | --- | --- | --- | ---: | --- |")
    for idx, issue in enumerate(issues, start=1):
        lines.append(
            "| "
            + " | ".join(
                [
                    f"ISSUE-{idx:04d}",
                    safe_cell(issue.category),
                    safe_cell(issue.sheet),
                    safe_cell(issue.row),
                    safe_cell(issue.source),
                    safe_cell(issue.source_iri),
                    safe_cell(issue.ttl_predicate),
                    safe_cell(issue.spreadsheet_relation),
                    safe_cell(issue.ttl_target),
                    safe_cell(issue.spreadsheet_target),
                    safe_cell(issue.ttl_line),
                    safe_cell(issue.recommended_action),
                ]
            )
            + " |"
        )
    lines.append("")
    lines.append("## Proposed Minimal Correction Plan")
    lines.append("")
    lines.append("### Proposed TTL Edits")
    lines.append(
        "- Do not edit authoritative generated `SSN2BFO.ttl` or frozen `legacy/SSN2BFO-pre-COMS.ttl`; this legacy audit is informational, and any accepted mapping change belongs in `mappings/SSN2BFO-COMS.xlsx`."
    )
    lines.append(
        "- For confirmed spreadsheet-governed rows, align the TTL predicate and named target IRIs with the spreadsheet axiom using the smallest possible axiom change."
    )
    lines.append(
        "- Treat rows involving unresolved prefixes or blank-node expression differences as human-review items, not mechanical edits."
    )
    lines.append("")
    lines.append("### Proposed Spreadsheet Edits")
    lines.append(
        "- For `missing_in_spreadsheet` findings, add spreadsheet rows only if the TTL assertion is intended to be governed by this source workbook."
    )
    lines.append(
        "- Add explicit status/review columns if maintainers want rejected, deferred, or provisional mappings to be machine-checkable."
    )
    lines.append(
        "- `sampling:` is accepted by this audit as an alias for the SOSA sample-relationship namespace; `sosa-rel:` is the preferred repo-facing alias."
    )
    lines.append("")
    lines.append("## Assumptions")
    lines.append("- The `OWL Axiom` column is the authoritative machine-comparison source for spreadsheet-governed mappings.")
    lines.append("- The workbook has no explicit status or review column in the inspected schema.")
    lines.append("- `sampling:` is accepted by this audit as an alias for the SOSA sample-relationship namespace; `sosa-rel:` is the preferred repo-facing alias for the same namespace.")
    lines.append("- Label-style `bfo:` and `cco:` spreadsheet tokens are resolved only when they map uniquely to imported labels.")
    path.write_text("\n".join(lines) + "\n", encoding="utf-8")


def main(argv: list[str] | None = None) -> int:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--ttl", required=True, type=Path)
    parser.add_argument("--spreadsheet", required=True, type=Path)
    parser.add_argument("--output-md", required=True, type=Path)
    parser.add_argument("--output-csv", required=True, type=Path)
    args = parser.parse_args(argv)

    repo_root = Path.cwd()
    if not args.ttl.exists():
        raise SystemExit(f"TTL file not found: {args.ttl}")
    if not args.spreadsheet.exists():
        raise SystemExit(f"Spreadsheet file not found: {args.spreadsheet}")

    label_index, _ = build_label_index(repo_root)
    resolver = Resolver(label_index)
    rows, schemas = parse_spreadsheet(args.spreadsheet, resolver)
    spreadsheet_sources = {row.source_iri for row in rows if row.source_iri}
    ttl_assertions, ignored = extract_ttl_assertions(args.ttl, spreadsheet_sources)
    issues, matched_ttl_keys = compare(rows, ttl_assertions)
    git_meta = git_metadata(repo_root)
    write_csv(args.output_csv, issues)
    write_markdown(
        args.output_md,
        ttl_path=args.ttl,
        spreadsheet_path=args.spreadsheet,
        exact_command=current_command(),
        portable_command_text=portable_command(),
        schemas=schemas,
        git_meta=git_meta,
        rows=rows,
        ttl_assertions=ttl_assertions,
        issues=issues,
        matched_ttl_keys=matched_ttl_keys,
        ignored=ignored,
    )
    print(f"inspected_sheets={len(schemas)}")
    print(f"spreadsheet_rows={len(rows)}")
    print(f"ttl_candidate_mapping_assertions={len(ttl_assertions)}")
    print(f"issues={len(issues)}")
    for category, count in sorted(Counter(issue.category for issue in issues).items()):
        print(f"{category}={count}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
